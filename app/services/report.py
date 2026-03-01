"""
app.services.report
-------------------
报告生成服务：支持 HTML、PDF、Markdown、Excel、JSON 多格式导出。
"""
import os
import json
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional

from jinja2 import Environment, FileSystemLoader
from app.models.task import ScanTask

TEMPLATE_DIR = "/app/app/templates"
OUTPUT_DIR = "/app/data/reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)


class ReportGenerator:
    """
    漏洞扫描报告生成器
    
    支持多种导出格式：
    - HTML: 网页格式，可直接在浏览器中查看
    - PDF: 文档格式，适合打印和存档
    - Markdown: 纯文本格式，可导入到其他工具
    - Excel: 表格格式，适合数据分析和筛选
    - JSON: 数据格式，适合程序处理和集成
    """
    
    def __init__(self):
        self.env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))

    def _get_summary(self, task: ScanTask) -> Dict[str, int]:
        """
        计算漏洞统计摘要
        
        Args:
            task: 扫描任务对象
            
        Returns:
            包含各等级漏洞数量的字典
        """
        summary = {
            "total": len(task.vulnerabilities),
            "critical": 0,
            "high": 0,
            "medium": 0,
            "low": 0,
            "info": 0
        }
        
        for vuln in task.vulnerabilities:
            sev = (vuln.severity or "").lower()
            if "critical" in sev:
                summary["critical"] += 1
            elif "high" in sev:
                summary["high"] += 1
            elif "medium" in sev:
                summary["medium"] += 1
            elif "low" in sev:
                summary["low"] += 1
            else:
                summary["info"] += 1
        
        return summary

    def generate_html(self, task: ScanTask, filename: str) -> str:
        """
        生成 HTML 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        template = self.env.get_template("report.html")
        summary = self._get_summary(task)
        html_content = template.render(task=task, vulns=task.vulnerabilities, summary=summary)
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        
        return output_path

    def generate_markdown(self, task: ScanTask, filename: str) -> str:
        """
        生成 Markdown 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        summary = self._get_summary(task)
        
        lines = []
        lines.append("# 🛡️ Aegis 漏洞检测报告")
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## 📋 扫描信息")
        lines.append("")
        lines.append(f"| 项目 | 值 |")
        lines.append("|------|-----|")
        lines.append(f"| 任务ID | {task.id} |")
        lines.append(f"| 目标URL | {task.target_url} |")
        lines.append(f"| 扫描时间 | {task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else 'N/A'} |")
        lines.append(f"| 扫描策略 | {task.scan_strategy or 'default'} |")
        lines.append(f"| 任务状态 | {task.status} |")
        lines.append("")
        
        # 漏洞统计
        lines.append("## 📊 漏洞统计")
        lines.append("")
        lines.append(f"| 风险等级 | 数量 | 状态 |")
        lines.append("|----------|------|------|")
        lines.append(f"| 🔴 严重 | {summary['critical']} | {'⚠️ 需立即修复' if summary['critical'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟠 高危 | {summary['high']} | {'⚠️ 需优先修复' if summary['high'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟡 中危 | {summary['medium']} | {'📋 建议修复' if summary['medium'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟢 低危 | {summary['low']} | {'📝 可选修复' if summary['low'] > 0 else '✅ 无'} |")
        lines.append(f"| ℹ️ 信息 | {summary['info']} | - |")
        lines.append(f"| **总计** | **{summary['total']}** | - |")
        lines.append("")
        
        # 漏洞详情
        lines.append("## 🔍 漏洞详情")
        lines.append("")
        
        if not task.vulnerabilities:
            lines.append("🎉 **太棒了！当前报告未发现漏洞。**")
        else:
            for idx, vuln in enumerate(task.vulnerabilities, start=1):
                severity_emoji = {
                    'critical': '🔴',
                    'high': '🟠',
                    'medium': '🟡',
                    'low': '🟢',
                    'info': 'ℹ️'
                }.get((vuln.severity or 'info').lower(), 'ℹ️')
                
                lines.append(f"### {idx}. {severity_emoji} [{vuln.severity.upper() if vuln.severity else 'N/A'}] {vuln.vuln_name or vuln.vuln_type or '未知漏洞'}")
                lines.append("")
                
                # 基本信息表格
                lines.append("| 属性 | 值 |")
                lines.append("|------|-----|")
                lines.append(f"| 漏洞类型 | {vuln.vuln_type or 'N/A'} |")
                lines.append(f"| 风险等级 | {vuln.severity or 'N/A'} |")
                lines.append(f"| 触发URL | `{vuln.url or 'N/A'}` |")
                lines.append(f"| 注入参数 | `{vuln.parameter or 'N/A'}` |")
                lines.append(f"| HTTP方法 | `{vuln.method or 'N/A'}` |")
                if vuln.cvss_score:
                    lines.append(f"| CVSS评分 | {vuln.cvss_score}/10 |")
                lines.append("")
                
                if vuln.description:
                    lines.append("**📝 漏洞描述**")
                    lines.append("")
                    lines.append(vuln.description)
                    lines.append("")
                
                # 攻击路径
                if vuln.attack_path:
                    lines.append("**🎯 模拟攻击路径**")
                    lines.append("")
                    if isinstance(vuln.attack_path, dict) and vuln.attack_path.get('steps'):
                        for step_idx, step in enumerate(vuln.attack_path['steps'], start=1):
                            method = step.get('method', 'GET')
                            url = step.get('url', '')
                            desc = step.get('description', '')
                            lines.append(f"{step_idx}. **[{method}]** `{url}`")
                            if desc:
                                lines.append(f"   - {desc}")
                    else:
                        # 简单路径
                        lines.append(f"1. **[{vuln.method or 'GET'}]** `{vuln.url or 'N/A'}`")
                        lines.append("   - 直接向目标发送恶意请求")
                    lines.append("")
                
                # 攻击载荷
                if vuln.payload:
                    lines.append("**💉 攻击载荷 (Payload)**")
                    lines.append("")
                    encoding_info = ""
                    if vuln.evidence and isinstance(vuln.evidence, dict):
                        enc = vuln.evidence.get('encoding_used', '')
                        if enc and enc != 'none':
                            encoding_info = f" (_编码类型: {enc}_)"
                    lines.append(f"```http{encoding_info}")
                    lines.append(vuln.payload)
                    lines.append("```")
                    lines.append("")
                
                # HTTP 请求详情
                if vuln.attack_path and isinstance(vuln.attack_path, dict) and vuln.attack_path.get('request'):
                    req = vuln.attack_path['request']
                    lines.append("**📡 HTTP 请求详情**")
                    lines.append("")
                    lines.append("```http")
                    lines.append(f"{req.get('method', 'GET')} {req.get('url', '')}")
                    if req.get('headers'):
                        for k, v in req['headers'].items():
                            lines.append(f"{k}: {v}")
                    if req.get('body'):
                        lines.append("")
                        lines.append(req['body'])
                    lines.append("```")
                    lines.append("")
                
                # 攻击证据
                if vuln.evidence:
                    lines.append("**🔎 攻击证据**")
                    lines.append("")
                    if isinstance(vuln.evidence, dict):
                        if vuln.evidence.get('matchers'):
                            lines.append("**匹配规则:**")
                            for m in vuln.evidence['matchers']:
                                lines.append(f"- 类型: {m.get('type', 'unknown')}")
                        if vuln.evidence.get('response_status'):
                            lines.append(f"- 响应状态码: {vuln.evidence['response_status']}")
                        if vuln.evidence.get('body_snippet'):
                            lines.append("")
                            lines.append("**响应片段:**")
                            lines.append("```")
                            lines.append(vuln.evidence['body_snippet'][:500])
                            lines.append("```")
                    else:
                        lines.append("```")
                        lines.append(str(vuln.evidence)[:500])
                        lines.append("```")
                    lines.append("")
                
                # 修复建议
                if vuln.remediation:
                    lines.append("**✅ 修复建议**")
                    lines.append("")
                    lines.append(f"> {vuln.remediation}")
                    lines.append("")
                
                lines.append("---")
                lines.append("")
        
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("*本报告由 Aegis Web应用程序漏洞检测系统自动生成*")
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        return output_path

    def generate_json(self, task: ScanTask, filename: str) -> str:
        """
        生成 JSON 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        summary = self._get_summary(task)
        
        vulnerabilities = []
        for vuln in task.vulnerabilities:
            # 安全获取可能不存在的新字段
            vuln_type = getattr(vuln, 'vuln_type', None)
            parameter = getattr(vuln, 'parameter', None)
            method = getattr(vuln, 'method', None)
            description = getattr(vuln, 'description', None)
            remediation = getattr(vuln, 'remediation', None)
            cvss_score = getattr(vuln, 'cvss_score', None)
            attack_path = getattr(vuln, 'attack_path', None)
            detected_at = getattr(vuln, 'detected_at', None)
            
            # 构建攻击路径信息
            attack_path_data = None
            if attack_path:
                attack_path_data = attack_path
            elif vuln.url:
                # 如果没有存储的攻击路径，基于基本信息构建
                attack_path_data = {
                    "steps": [
                        {
                            "step": 1,
                            "method": method or "GET",
                            "url": vuln.url,
                            "description": "直接向目标发送恶意请求"
                        }
                    ],
                    "request": {
                        "method": method or "GET",
                        "url": vuln.url,
                        "headers": {},
                        "body": vuln.payload if method == "POST" else None
                    }
                }
            
            vuln_data = {
                "id": vuln.id,
                "name": vuln.vuln_name,
                "type": vuln_type,
                "severity": vuln.severity,
                "cvss_score": cvss_score,
                "url": vuln.url,
                "parameter": parameter,
                "method": method,
                "description": description,
                "remediation": remediation,
                # 攻击路径 - 核心改进
                "attack_path": attack_path_data,
                # 攻击载荷 - 核心改进
                "payload": {
                    "raw": vuln.payload,
                    "encoded": vuln.payload,  # 兼容旧数据
                    "encoding_type": None,
                    "original_payload": None
                },
                # 攻击证据
                "evidence": vuln.evidence,
                "detected_at": detected_at.isoformat() if detected_at else None
            }
            
            # 从 evidence 中提取编码信息
            if vuln.evidence and isinstance(vuln.evidence, dict):
                vuln_data["payload"]["encoding_type"] = vuln.evidence.get("encoding_used")
                vuln_data["payload"]["mutation_type"] = vuln.evidence.get("mutation_type")
                # 如果有原始 payload 信息
                if vuln.evidence.get("request") and isinstance(vuln.evidence["request"], dict):
                    vuln_data["payload"]["original_payload"] = vuln.evidence["request"].get("payload_original")
            
            vulnerabilities.append(vuln_data)
        
        report = {
            "report_info": {
                "task_id": task.id,
                "target_url": task.target_url,
                "scan_strategy": task.scan_strategy,
                "status": task.status,
                "scan_time": task.updated_at.isoformat() if task.updated_at else None,
                "generated_at": datetime.now().isoformat(),
                "generator": "Aegis Web应用程序漏洞检测系统"
            },
            "summary": summary,
            "vulnerabilities": vulnerabilities,
            # 添加攻击模拟摘要
            "attack_simulation": {
                "total_payloads": len([v for v in task.vulnerabilities if v.payload]),
                "total_attack_paths": len([v for v in task.vulnerabilities if v.attack_path]),
                "encoding_types_used": list(set([
                    v.evidence.get("encoding_used") 
                    for v in task.vulnerabilities 
                    if v.evidence and isinstance(v.evidence, dict) and v.evidence.get("encoding_used")
                ]))
            }
        }
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        return output_path

    def generate_excel(self, task: ScanTask, filename: str) -> str:
        """
        生成 Excel 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # 定义样式
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2d3343", end_color="2d3343", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        summary = self._get_summary(task)
        
        # ========== 概览工作表 ==========
        ws_overview = wb.active
        ws_overview.title = "扫描概览"
        
        # 标题
        ws_overview.merge_cells('A1:D1')
        ws_overview['A1'] = "🛡️ Aegis Web应用程序漏洞检测报告"
        ws_overview['A1'].font = Font(bold=True, size=18)
        ws_overview['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        ws_overview['A3'] = "📋 基本信息"
        ws_overview['A3'].font = subheader_font
        ws_overview.merge_cells('A3:D3')
        ws_overview['A3'].fill = subheader_fill
        
        basic_info = [
            ("任务ID", task.id),
            ("目标URL", task.target_url),
            ("扫描时间", task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else "N/A"),
            ("扫描策略", task.scan_strategy or "default"),
            ("任务状态", task.status),
        ]
        
        for idx, (label, value) in enumerate(basic_info, start=4):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = str(value)
            ws_overview.merge_cells(f'B{idx}:D{idx}')
        
        # 漏洞统计
        row = len(basic_info) + 5
        ws_overview[f'A{row}'] = "📊 漏洞统计"
        ws_overview[f'A{row}'].font = subheader_font
        ws_overview.merge_cells(f'A{row}:D{row}')
        ws_overview[f'A{row}'].fill = subheader_fill
        
        vuln_stats = [
            ("总漏洞数", summary["total"]),
            ("🔴 严重", summary["critical"]),
            ("🟠 高危", summary["high"]),
            ("🟡 中危", summary["medium"]),
            ("🟢 低危", summary["low"]),
            ("ℹ️ 信息", summary["info"]),
        ]
        
        # 风险等级颜色
        risk_colors = {
            "🔴 严重": "dc3545",
            "🟠 高危": "fd7e14",
            "🟡 中危": "ffc107",
            "🟢 低危": "28a745",
            "ℹ️ 信息": "17a2b8",
        }
        
        for idx, (label, value) in enumerate(vuln_stats, start=row+1):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = value
            if label in risk_colors:
                ws_overview[f'A{idx}'].fill = PatternFill(start_color=risk_colors[label], end_color=risk_colors[label], fill_type="solid")
                ws_overview[f'A{idx}'].font = Font(bold=True, color="FFFFFF")
        
        # 设置列宽
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 30
        ws_overview.column_dimensions['C'].width = 20
        ws_overview.column_dimensions['D'].width = 20
        
        # ========== 漏洞详情工作表 ==========
        ws_vulns = wb.create_sheet("漏洞详情")
        
        # 表头 - 扩展包含攻击载荷
        headers = ["序号", "漏洞名称", "风险等级", "漏洞类型", "URL", "参数", "HTTP方法", "CVSS评分", "攻击载荷", "漏洞描述", "修复建议"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_vulns.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 填充漏洞数据
        for idx, vuln in enumerate(task.vulnerabilities, start=1):
            row = idx + 1
            col = 1
            
            ws_vulns.cell(row=row, column=col, value=idx).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.vuln_name or vuln.vuln_type or "未知").border = border
            col += 1
            
            # 风险等级单元格
            risk_cell = ws_vulns.cell(row=row, column=col, value=(vuln.severity or "info").upper())
            risk_cell.border = border
            risk_colors_cell = {
                'critical': "dc3545",
                'high': "fd7e14",
                'medium': "ffc107",
                'low': "28a745",
                'info': "17a2b8",
            }
            sev_lower = (vuln.severity or "info").lower()
            if sev_lower in risk_colors_cell:
                risk_cell.fill = PatternFill(start_color=risk_colors_cell[sev_lower], 
                                            end_color=risk_colors_cell[sev_lower], 
                                            fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            col += 1
            
            ws_vulns.cell(row=row, column=col, value=vuln.vuln_type or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.url or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.parameter or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.method or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.cvss_score or "N/A").border = border
            col += 1
            # 攻击载荷列
            payload_cell = ws_vulns.cell(row=row, column=col, value=(vuln.payload or "N/A")[:200])
            payload_cell.border = border
            payload_cell.font = Font(name='Consolas', size=9)
            col += 1
            ws_vulns.cell(row=row, column=col, value=(vuln.description or "N/A")[:100]).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=(vuln.remediation or "N/A")[:100]).border = border
        
        # 设置列宽
        column_widths = [8, 30, 10, 15, 40, 15, 10, 10, 50, 40, 40]
        for col, width in enumerate(column_widths, start=1):
            ws_vulns.column_dimensions[get_column_letter(col)].width = width
        
        # ========== 攻击路径详情工作表 ==========
        ws_attack = wb.create_sheet("攻击路径详情")
        
        # 表头
        attack_headers = ["漏洞ID", "漏洞名称", "攻击步骤", "HTTP方法", "攻击URL", "载荷", "编码类型", "请求详情"]
        
        for col, header in enumerate(attack_headers, start=1):
            cell = ws_attack.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 填充攻击路径数据
        row = 2
        for vuln in task.vulnerabilities:
            if vuln.attack_path and isinstance(vuln.attack_path, dict):
                steps = vuln.attack_path.get('steps', [])
                if steps:
                    for step in steps:
                        ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                        ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                        ws_attack.cell(row=row, column=3, value=step.get('step', 1)).border = border
                        ws_attack.cell(row=row, column=4, value=step.get('method', 'GET')).border = border
                        ws_attack.cell(row=row, column=5, value=step.get('url', vuln.url or 'N/A')).border = border
                        ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                        
                        # 编码类型
                        enc_type = "原始"
                        if vuln.evidence and isinstance(vuln.evidence, dict):
                            enc = vuln.evidence.get('encoding_used', '')
                            if enc and enc != 'none':
                                enc_type = enc
                        ws_attack.cell(row=row, column=7, value=enc_type).border = border
                        
                        # 请求详情
                        req_detail = ""
                        if vuln.attack_path.get('request'):
                            req = vuln.attack_path['request']
                            req_detail = f"{req.get('method', 'GET')} {req.get('url', '')}"
                        ws_attack.cell(row=row, column=8, value=req_detail).border = border
                        
                        row += 1
                else:
                    # 简单路径
                    ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                    ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                    ws_attack.cell(row=row, column=3, value=1).border = border
                    ws_attack.cell(row=row, column=4, value=vuln.method or 'GET').border = border
                    ws_attack.cell(row=row, column=5, value=vuln.url or 'N/A').border = border
                    ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                    ws_attack.cell(row=row, column=7, value="原始").border = border
                    ws_attack.cell(row=row, column=8, value=f"{vuln.method or 'GET'} {vuln.url or 'N/A'}").border = border
                    row += 1
            elif vuln.url:
                # 无攻击路径但有URL
                ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                ws_attack.cell(row=row, column=3, value=1).border = border
                ws_attack.cell(row=row, column=4, value=vuln.method or 'GET').border = border
                ws_attack.cell(row=row, column=5, value=vuln.url).border = border
                ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                ws_attack.cell(row=row, column=7, value="原始").border = border
                ws_attack.cell(row=row, column=8, value=f"{vuln.method or 'GET'} {vuln.url}").border = border
                row += 1
        
        # 设置列宽
        attack_widths = [10, 30, 10, 10, 50, 40, 15, 50]
        for col, width in enumerate(attack_widths, start=1):
            ws_attack.column_dimensions[get_column_letter(col)].width = width
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        wb.save(output_path)
        
        return output_path

    def generate_pdf(self, task: ScanTask, filename: str) -> str:
        """
        生成 PDF 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import platform
        
        # 注册中文字体
        try:
            if platform.system() == 'Windows':
                font_path = r'C:\Windows\Fonts\simsun.ttc'
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('SimSun', font_path, subfontIndex=0))
            elif platform.system() == 'Linux':
                for font_path in ['/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
                                 '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf']:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('SimSun', font_path))
                        break
            elif platform.system() == 'Darwin':
                for font_path in ['/System/Library/Fonts/PingFang.ttc']:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('SimSun', font_path))
                        break
        except Exception:
            pass
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        
        styles = getSampleStyleSheet()
        
        # 自定义样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue,
            fontName='SimSun'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            textColor=colors.darkblue,
            fontName='SimSun'
        )
        
        normal_style = styles['Normal']
        normal_style.fontSize = 10
        normal_style.fontName = 'SimSun'
        
        summary = self._get_summary(task)
        
        # 构建PDF内容
        story = []
        
        # 标题
        story.append(Paragraph("Web应用程序漏洞检测报告", title_style))
        story.append(Spacer(1, 20))
        
        # 基本信息表格
        basic_info_data = [
            ['任务ID', str(task.id)],
            ['目标URL', task.target_url],
            ['扫描时间', task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else "N/A"],
            ['任务状态', task.status],
        ]
        
        basic_table = Table(basic_info_data, colWidths=[2*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(basic_table)
        story.append(Spacer(1, 20))
        
        # 漏洞统计
        story.append(Paragraph("漏洞统计", heading_style))
        
        stats_data = [
            ['总漏洞数', str(summary['total'])],
            ['严重漏洞', str(summary['critical'])],
            ['高危漏洞', str(summary['high'])],
            ['中危漏洞', str(summary['medium'])],
            ['低危漏洞', str(summary['low'])],
            ['信息', str(summary['info'])],
        ]
        
        stats_table = Table(stats_data, colWidths=[2*inch, 4*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # 漏洞详情
        if task.vulnerabilities:
            story.append(Paragraph("漏洞详情", heading_style))
            
            for idx, vuln in enumerate(task.vulnerabilities, start=1):
                vuln_title = f"{idx}. [{vuln.severity.upper() if vuln.severity else 'N/A'}] {vuln.title or vuln.vuln_type or '未知漏洞'}"
                story.append(Paragraph(vuln_title, ParagraphStyle(
                    'VulnTitle',
                    parent=styles['Heading3'],
                    fontSize=12,
                    spaceAfter=10,
                    fontName='SimSun'
                )))
                
                vuln_data = [
                    ['漏洞类型', vuln.vuln_type or 'N/A'],
                    ['风险等级', vuln.severity or 'N/A'],
                    ['URL', vuln.url or 'N/A'],
                    ['参数', vuln.parameter or 'N/A'],
                    ['漏洞描述', (vuln.description or 'N/A')[:100]],
                    ['修复建议', (vuln.remediation or 'N/A')[:100]],
                ]
                
                vuln_table = Table(vuln_data, colWidths=[1.5*inch, 4.5*inch])
                vuln_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(vuln_table)
                story.append(Spacer(1, 15))
        else:
            story.append(Paragraph("当前报告未发现漏洞。", normal_style))
        
        doc.build(story)
        
        return output_path

    def generate(self, task: ScanTask, format: str = "html") -> tuple[str, str]:
        """
        根据指定格式生成报告
        
        Args:
            task: 扫描任务对象
            format: 导出格式 (html/pdf/markdown/excel/json)
            
        Returns:
            (文件路径, 文件名) 元组
        """
        format = format.lower()
        
        # 文件扩展名映射
        extensions = {
            "html": ".html",
            "pdf": ".pdf",
            "markdown": ".md",
            "excel": ".xlsx",
            "json": ".json"
        }
        
        # MIME 类型映射
        mime_types = {
            "html": "text/html",
            "pdf": "application/pdf",
            "markdown": "text/markdown",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json"
        }
        
        if format not in extensions:
            format = "html"
        
        filename = f"report_{task.id}{extensions[format]}"
        
        # 根据格式调用相应的生成方法
        generators = {
            "html": self.generate_html,
            "markdown": self.generate_markdown,
            "json": self.generate_json,
            "excel": self.generate_excel,
            "pdf": self.generate_pdf
        }
        
        file_path = generators[format](task, filename)
        
        return file_path, filename