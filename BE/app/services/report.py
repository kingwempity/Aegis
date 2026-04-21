"""
app.services.report
-------------------
报告生成服务：支持 HTML、PDF、Markdown、Excel、JSON 多格式导出。
"""
import os
import json
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional

from jinja2 import Environment, FileSystemLoader
from app.models.task import ScanTask

TEMPLATE_DIR = "/app/BE/app/templates"
OUTPUT_DIR = "/app/data/reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)


class ReportGenerator:
    """
    漏洞扫描报告生成器
    
    支持多种导出格式：
    - HTML: 网页格式，可直接在浏览器中查看
    - PDF: 文档格式，适合打印和存档
    - Markdown: 纯文本格式，可导入到其他工具
    - Excel: 表格格式，适合数据分析和筛选
    - JSON: 数据格式，适合程序处理和集成
    """
    
    def __init__(self):
        self.env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))

    def _get_summary(self, task: ScanTask) -> Dict[str, int]:
        """
        计算漏洞统计摘要
        
        Args:
            task: 扫描任务对象
            
        Returns:
            包含各等级漏洞数量的字典
        """
        summary = {
            "total": len(task.vulnerabilities),
            "critical": 0,
            "high": 0,
            "medium": 0,
            "low": 0,
            "info": 0
        }
        
        for vuln in task.vulnerabilities:
            sev = (vuln.severity or "").lower()
            if "critical" in sev:
                summary["critical"] += 1
            elif "high" in sev:
                summary["high"] += 1
            elif "medium" in sev:
                summary["medium"] += 1
            elif "low" in sev:
                summary["low"] += 1
            else:
                summary["info"] += 1
        
        return summary

    def _get_status_display(self, status: str) -> str:
        """
        将任务状态转换为中文显示
        
        Args:
            status: 任务状态代码
            
        Returns:
            中文状态描述
        """
        status_map = {
            "pending": "等待中",
            "running": "扫描中",
            "completed": "已完成",
            "failed": "失败",
            "cancelled": "已取消",
        }
        return status_map.get(status, status) or "未知"

    def _get_severity_display(self, severity: str) -> str:
        """
        将风险等级转换为中文显示
        
        Args:
            severity: 风险等级代码
            
        Returns:
            中文风险等级描述
        """
        if not severity:
            return "未知"
        
        sev = severity.lower()
        severity_map = {
            "critical": "严重",
            "high": "高危",
            "medium": "中危",
            "low": "低危",
            "info": "信息",
        }
        return severity_map.get(sev, severity)

    def generate_html(self, task: ScanTask, filename: str) -> str:
        """
        生成 HTML 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        template = self.env.get_template("report.html")
        summary = self._get_summary(task)
        html_content = template.render(task=task, vulns=task.vulnerabilities, summary=summary)
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        
        return output_path

    def generate_markdown(self, task: ScanTask, filename: str) -> str:
        """
        生成 Markdown 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        summary = self._get_summary(task)
        
        lines = []
        lines.append("# 🛡️ Aegis 漏洞检测报告")
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## 📋 扫描信息")
        lines.append("")
        lines.append(f"| 项目 | 值 |")
        lines.append("|------|-----|")
        lines.append(f"| 任务ID | {task.id} |")
        lines.append(f"| 目标URL | {task.target_url} |")
        lines.append(f"| 扫描时间 | {task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else 'N/A'} |")
        lines.append(f"| 扫描策略 | {task.scan_strategy or 'default'} |")
        lines.append(f"| 任务状态 | {task.status} |")
        lines.append("")
        
        # 漏洞统计
        lines.append("## 📊 漏洞统计")
        lines.append("")
        lines.append(f"| 风险等级 | 数量 | 状态 |")
        lines.append("|----------|------|------|")
        lines.append(f"| 🔴 严重 | {summary['critical']} | {'⚠️ 需立即修复' if summary['critical'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟠 高危 | {summary['high']} | {'⚠️ 需优先修复' if summary['high'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟡 中危 | {summary['medium']} | {'📋 建议修复' if summary['medium'] > 0 else '✅ 无'} |")
        lines.append(f"| 🟢 低危 | {summary['low']} | {'📝 可选修复' if summary['low'] > 0 else '✅ 无'} |")
        lines.append(f"| ℹ️ 信息 | {summary['info']} | - |")
        lines.append(f"| **总计** | **{summary['total']}** | - |")
        lines.append("")
        
        # 漏洞详情
        lines.append("## 🔍 漏洞详情")
        lines.append("")
        
        if not task.vulnerabilities:
            lines.append("🎉 **太棒了！当前报告未发现漏洞。**")
        else:
            for idx, vuln in enumerate(task.vulnerabilities, start=1):
                severity_emoji = {
                    'critical': '🔴',
                    'high': '🟠',
                    'medium': '🟡',
                    'low': '🟢',
                    'info': 'ℹ️'
                }.get((vuln.severity or 'info').lower(), 'ℹ️')
                
                lines.append(f"### {idx}. {severity_emoji} [{vuln.severity.upper() if vuln.severity else 'N/A'}] {vuln.vuln_name or vuln.vuln_type or '未知漏洞'}")
                lines.append("")
                
                # 基本信息表格
                lines.append("| 属性 | 值 |")
                lines.append("|------|-----|")
                lines.append(f"| 漏洞类型 | {vuln.vuln_type or 'N/A'} |")
                lines.append(f"| 风险等级 | {vuln.severity or 'N/A'} |")
                lines.append(f"| 触发URL | `{vuln.url or 'N/A'}` |")
                lines.append(f"| 注入参数 | `{vuln.parameter or 'N/A'}` |")
                lines.append(f"| HTTP方法 | `{vuln.method or 'N/A'}` |")
                if vuln.cvss_score:
                    lines.append(f"| CVSS评分 | {vuln.cvss_score}/10 |")
                lines.append("")
                
                if vuln.description:
                    lines.append("**📝 漏洞描述**")
                    lines.append("")
                    lines.append(vuln.description)
                    lines.append("")
                
                # 攻击路径
                if vuln.attack_path:
                    lines.append("**🎯 模拟攻击路径**")
                    lines.append("")
                    if isinstance(vuln.attack_path, dict) and vuln.attack_path.get('steps'):
                        for step_idx, step in enumerate(vuln.attack_path['steps'], start=1):
                            method = step.get('method', 'GET')
                            url = step.get('url', '')
                            desc = step.get('description', '')
                            lines.append(f"{step_idx}. **[{method}]** `{url}`")
                            if desc:
                                lines.append(f"   - {desc}")
                    else:
                        # 简单路径
                        lines.append(f"1. **[{vuln.method or 'GET'}]** `{vuln.url or 'N/A'}`")
                        lines.append("   - 直接向目标发送恶意请求")
                    lines.append("")
                
                # 攻击载荷
                if vuln.payload:
                    lines.append("**💉 攻击载荷 (Payload)**")
                    lines.append("")
                    encoding_info = ""
                    if vuln.evidence and isinstance(vuln.evidence, dict):
                        enc = vuln.evidence.get('encoding_used', '')
                        if enc and enc != 'none':
                            encoding_info = f" (_编码类型: {enc}_)"
                    lines.append(f"```http{encoding_info}")
                    lines.append(vuln.payload)
                    lines.append("```")
                    lines.append("")
                
                # HTTP 请求详情
                if vuln.attack_path and isinstance(vuln.attack_path, dict) and vuln.attack_path.get('request'):
                    req = vuln.attack_path['request']
                    lines.append("**📡 HTTP 请求详情**")
                    lines.append("")
                    lines.append("```http")
                    lines.append(f"{req.get('method', 'GET')} {req.get('url', '')}")
                    if req.get('headers'):
                        for k, v in req['headers'].items():
                            lines.append(f"{k}: {v}")
                    if req.get('body'):
                        lines.append("")
                        lines.append(req['body'])
                    lines.append("```")
                    lines.append("")
                
                # 攻击证据
                if vuln.evidence:
                    lines.append("**🔎 攻击证据**")
                    lines.append("")
                    if isinstance(vuln.evidence, dict):
                        if vuln.evidence.get('matchers'):
                            lines.append("**匹配规则:**")
                            for m in vuln.evidence['matchers']:
                                lines.append(f"- 类型: {m.get('type', 'unknown')}")
                        if vuln.evidence.get('response_status'):
                            lines.append(f"- 响应状态码: {vuln.evidence['response_status']}")
                        if vuln.evidence.get('body_snippet'):
                            lines.append("")
                            lines.append("**响应片段:**")
                            lines.append("```")
                            lines.append(vuln.evidence['body_snippet'][:500])
                            lines.append("```")
                    else:
                        lines.append("```")
                        lines.append(str(vuln.evidence)[:500])
                        lines.append("```")
                    lines.append("")
                
                # 修复建议
                if vuln.remediation:
                    lines.append("**✅ 修复建议**")
                    lines.append("")
                    lines.append(f"> {vuln.remediation}")
                    lines.append("")
                
                lines.append("---")
                lines.append("")
        
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("*本报告由 Aegis Web应用程序漏洞检测系统自动生成*")
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        return output_path

    def generate_json(self, task: ScanTask, filename: str) -> str:
        """
        生成 JSON 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        summary = self._get_summary(task)
        
        vulnerabilities = []
        for vuln in task.vulnerabilities:
            # 安全获取可能不存在的新字段
            vuln_type = getattr(vuln, 'vuln_type', None)
            parameter = getattr(vuln, 'parameter', None)
            method = getattr(vuln, 'method', None)
            description = getattr(vuln, 'description', None)
            remediation = getattr(vuln, 'remediation', None)
            cvss_score = getattr(vuln, 'cvss_score', None)
            attack_path = getattr(vuln, 'attack_path', None)
            detected_at = getattr(vuln, 'detected_at', None)
            
            # 构建攻击路径信息
            attack_path_data = None
            if attack_path:
                attack_path_data = attack_path
            elif vuln.url:
                # 如果没有存储的攻击路径，基于基本信息构建
                attack_path_data = {
                    "steps": [
                        {
                            "step": 1,
                            "method": method or "GET",
                            "url": vuln.url,
                            "description": "直接向目标发送恶意请求"
                        }
                    ],
                    "request": {
                        "method": method or "GET",
                        "url": vuln.url,
                        "headers": {},
                        "body": vuln.payload if method == "POST" else None
                    }
                }
            
            vuln_data = {
                "id": vuln.id,
                "name": vuln.vuln_name,
                "type": vuln_type,
                "severity": vuln.severity,
                "cvss_score": cvss_score,
                "url": vuln.url,
                "parameter": parameter,
                "method": method,
                "description": description,
                "remediation": remediation,
                # 攻击路径 - 核心改进
                "attack_path": attack_path_data,
                # 攻击载荷 - 核心改进
                "payload": {
                    "raw": vuln.payload,
                    "encoded": vuln.payload,  # 兼容旧数据
                    "encoding_type": None,
                    "original_payload": None
                },
                # 攻击证据
                "evidence": vuln.evidence,
                "detected_at": detected_at.isoformat() if detected_at else None
            }
            
            # 从 evidence 中提取编码信息
            if vuln.evidence and isinstance(vuln.evidence, dict):
                vuln_data["payload"]["encoding_type"] = vuln.evidence.get("encoding_used")
                vuln_data["payload"]["mutation_type"] = vuln.evidence.get("mutation_type")
                # 如果有原始 payload 信息
                if vuln.evidence.get("request") and isinstance(vuln.evidence["request"], dict):
                    vuln_data["payload"]["original_payload"] = vuln.evidence["request"].get("payload_original")
            
            vulnerabilities.append(vuln_data)
        
        report = {
            "report_info": {
                "task_id": task.id,
                "target_url": task.target_url,
                "scan_strategy": task.scan_strategy,
                "status": task.status,
                "scan_time": task.updated_at.isoformat() if task.updated_at else None,
                "generated_at": datetime.now().isoformat(),
                "generator": "Aegis Web应用程序漏洞检测系统"
            },
            "summary": summary,
            "vulnerabilities": vulnerabilities,
            # 添加攻击模拟摘要
            "attack_simulation": {
                "total_payloads": len([v for v in task.vulnerabilities if v.payload]),
                "total_attack_paths": len([v for v in task.vulnerabilities if v.attack_path]),
                "encoding_types_used": list(set([
                    v.evidence.get("encoding_used") 
                    for v in task.vulnerabilities 
                    if v.evidence and isinstance(v.evidence, dict) and v.evidence.get("encoding_used")
                ]))
            }
        }
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        return output_path

    def generate_excel(self, task: ScanTask, filename: str) -> str:
        """
        生成 Excel 报告并返回文件路径
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # 定义样式
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2d3343", end_color="2d3343", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        summary = self._get_summary(task)
        
        # ========== 概览工作表 ==========
        ws_overview = wb.active
        ws_overview.title = "扫描概览"
        
        # 标题
        ws_overview.merge_cells('A1:D1')
        ws_overview['A1'] = "🛡️ Aegis Web应用程序漏洞检测报告"
        ws_overview['A1'].font = Font(bold=True, size=18)
        ws_overview['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        ws_overview['A3'] = "📋 基本信息"
        ws_overview['A3'].font = subheader_font
        ws_overview.merge_cells('A3:D3')
        ws_overview['A3'].fill = subheader_fill
        
        basic_info = [
            ("任务ID", task.id),
            ("目标URL", task.target_url),
            ("扫描时间", task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else "N/A"),
            ("扫描策略", task.scan_strategy or "default"),
            ("任务状态", task.status),
        ]
        
        for idx, (label, value) in enumerate(basic_info, start=4):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = str(value)
            ws_overview.merge_cells(f'B{idx}:D{idx}')
        
        # 漏洞统计
        row = len(basic_info) + 5
        ws_overview[f'A{row}'] = "📊 漏洞统计"
        ws_overview[f'A{row}'].font = subheader_font
        ws_overview.merge_cells(f'A{row}:D{row}')
        ws_overview[f'A{row}'].fill = subheader_fill
        
        vuln_stats = [
            ("总漏洞数", summary["total"]),
            ("🔴 严重", summary["critical"]),
            ("🟠 高危", summary["high"]),
            ("🟡 中危", summary["medium"]),
            ("🟢 低危", summary["low"]),
            ("ℹ️ 信息", summary["info"]),
        ]
        
        # 风险等级颜色
        risk_colors = {
            "🔴 严重": "dc3545",
            "🟠 高危": "fd7e14",
            "🟡 中危": "ffc107",
            "🟢 低危": "28a745",
            "ℹ️ 信息": "17a2b8",
        }
        
        for idx, (label, value) in enumerate(vuln_stats, start=row+1):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = value
            if label in risk_colors:
                ws_overview[f'A{idx}'].fill = PatternFill(start_color=risk_colors[label], end_color=risk_colors[label], fill_type="solid")
                ws_overview[f'A{idx}'].font = Font(bold=True, color="FFFFFF")
        
        # 设置列宽
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 30
        ws_overview.column_dimensions['C'].width = 20
        ws_overview.column_dimensions['D'].width = 20
        
        # ========== 漏洞详情工作表 ==========
        ws_vulns = wb.create_sheet("漏洞详情")
        
        # 表头 - 扩展包含攻击载荷
        headers = ["序号", "漏洞名称", "风险等级", "漏洞类型", "URL", "参数", "HTTP方法", "CVSS评分", "攻击载荷", "漏洞描述", "修复建议"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_vulns.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 填充漏洞数据
        for idx, vuln in enumerate(task.vulnerabilities, start=1):
            row = idx + 1
            col = 1
            
            ws_vulns.cell(row=row, column=col, value=idx).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.vuln_name or vuln.vuln_type or "未知").border = border
            col += 1
            
            # 风险等级单元格
            risk_cell = ws_vulns.cell(row=row, column=col, value=(vuln.severity or "info").upper())
            risk_cell.border = border
            risk_colors_cell = {
                'critical': "dc3545",
                'high': "fd7e14",
                'medium': "ffc107",
                'low': "28a745",
                'info': "17a2b8",
            }
            sev_lower = (vuln.severity or "info").lower()
            if sev_lower in risk_colors_cell:
                risk_cell.fill = PatternFill(start_color=risk_colors_cell[sev_lower], 
                                            end_color=risk_colors_cell[sev_lower], 
                                            fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            col += 1
            
            ws_vulns.cell(row=row, column=col, value=vuln.vuln_type or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.url or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.parameter or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.method or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.cvss_score or "N/A").border = border
            col += 1
            # 攻击载荷列
            payload_cell = ws_vulns.cell(row=row, column=col, value=(vuln.payload or "N/A")[:200])
            payload_cell.border = border
            payload_cell.font = Font(name='Consolas', size=9)
            col += 1
            ws_vulns.cell(row=row, column=col, value=(vuln.description or "N/A")[:100]).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=(vuln.remediation or "N/A")[:100]).border = border
        
        # 设置列宽
        column_widths = [8, 30, 10, 15, 40, 15, 10, 10, 50, 40, 40]
        for col, width in enumerate(column_widths, start=1):
            ws_vulns.column_dimensions[get_column_letter(col)].width = width
        
        # ========== 攻击路径详情工作表 ==========
        ws_attack = wb.create_sheet("攻击路径详情")
        
        # 表头
        attack_headers = ["漏洞ID", "漏洞名称", "攻击步骤", "HTTP方法", "攻击URL", "载荷", "编码类型", "请求详情"]
        
        for col, header in enumerate(attack_headers, start=1):
            cell = ws_attack.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 填充攻击路径数据
        row = 2
        for vuln in task.vulnerabilities:
            if vuln.attack_path and isinstance(vuln.attack_path, dict):
                steps = vuln.attack_path.get('steps', [])
                if steps:
                    for step in steps:
                        ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                        ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                        ws_attack.cell(row=row, column=3, value=step.get('step', 1)).border = border
                        ws_attack.cell(row=row, column=4, value=step.get('method', 'GET')).border = border
                        ws_attack.cell(row=row, column=5, value=step.get('url', vuln.url or 'N/A')).border = border
                        ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                        
                        # 编码类型
                        enc_type = "原始"
                        if vuln.evidence and isinstance(vuln.evidence, dict):
                            enc = vuln.evidence.get('encoding_used', '')
                            if enc and enc != 'none':
                                enc_type = enc
                        ws_attack.cell(row=row, column=7, value=enc_type).border = border
                        
                        # 请求详情
                        req_detail = ""
                        if vuln.attack_path.get('request'):
                            req = vuln.attack_path['request']
                            req_detail = f"{req.get('method', 'GET')} {req.get('url', '')}"
                        ws_attack.cell(row=row, column=8, value=req_detail).border = border
                        
                        row += 1
                else:
                    # 简单路径
                    ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                    ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                    ws_attack.cell(row=row, column=3, value=1).border = border
                    ws_attack.cell(row=row, column=4, value=vuln.method or 'GET').border = border
                    ws_attack.cell(row=row, column=5, value=vuln.url or 'N/A').border = border
                    ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                    ws_attack.cell(row=row, column=7, value="原始").border = border
                    ws_attack.cell(row=row, column=8, value=f"{vuln.method or 'GET'} {vuln.url or 'N/A'}").border = border
                    row += 1
            elif vuln.url:
                # 无攻击路径但有URL
                ws_attack.cell(row=row, column=1, value=vuln.id).border = border
                ws_attack.cell(row=row, column=2, value=vuln.vuln_name or "未知").border = border
                ws_attack.cell(row=row, column=3, value=1).border = border
                ws_attack.cell(row=row, column=4, value=vuln.method or 'GET').border = border
                ws_attack.cell(row=row, column=5, value=vuln.url).border = border
                ws_attack.cell(row=row, column=6, value=(vuln.payload or 'N/A')[:100]).border = border
                ws_attack.cell(row=row, column=7, value="原始").border = border
                ws_attack.cell(row=row, column=8, value=f"{vuln.method or 'GET'} {vuln.url}").border = border
                row += 1
        
        # 设置列宽
        attack_widths = [10, 30, 10, 10, 50, 40, 15, 50]
        for col, width in enumerate(attack_widths, start=1):
            ws_attack.column_dimensions[get_column_letter(col)].width = width
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        wb.save(output_path)
        
        return output_path

    def generate_pdf(self, task: ScanTask, filename: str) -> str:
        """
        生成 PDF 报告并返回文件路径 (Acunetix风格)
        
        Args:
            task: 扫描任务对象
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.graphics import renderPDF
        import platform
        import logging
        
        logger = logging.getLogger(__name__)
        
        FONT_NAME = 'ChineseFont'
        
        def register_chinese_font():
            """注册中文字体，按优先级尝试多种方式和字体"""
            
            def try_register(font_path, font_name, subfont_index=None):
                """尝试注册单个字体"""
                if not os.path.exists(font_path):
                    return False, f"文件不存在: {font_path}"
                try:
                    if subfont_index is not None:
                        pdfmetrics.registerFont(TTFont(font_name, font_path, subfontIndex=subfont_index))
                    else:
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                    
                    test_style = ParagraphStyle('Test', fontName=font_name, fontSize=10)
                    from reportlab.platypus import Paragraph
                    from io import BytesIO
                    from reportlab.pdfgen import canvas
                    buffer = BytesIO()
                    c = canvas.Canvas(buffer, pagesize=(100, 100))
                    c.setFont(font_name, 10)
                    c.drawString(10, 50, "测试中文")
                    c.save()
                    return True, f"成功: {font_path}"
                except Exception as e:
                    return False, f"{type(e).__name__}: {str(e)}"
            
            import subprocess
            
            system = platform.system()
            logger.info(f"当前操作系统: {system}, 尝试注册中文字体...")
            
            candidates = []
            
            if system == 'Windows':
                candidates = [
                    (r'C:\Windows\Fonts\simsun.ttc', 'SimSun', [0]),
                    (r'C:\Windows\Fonts\msyh.ttc', 'MsYH', [0]),
                    (r'C:\Windows\Fonts\simhei.ttf', 'SimHei', [None]),
                    (r'C:\Windows\Fonts\msyhbd.ttc', 'MsYH-Bold', [0]),
                    (r'C:\Windows\Fonts\simsunb.ttf', 'SimSun-Bold', [None]),
                ]
            elif system == 'Linux':
                try:
                    result = subprocess.run(['fc-list', ':lang=zh', 'file'], 
                                          capture_output=True, text=True, timeout=5)
                    if result.returncode == 0 and result.stdout.strip():
                        fonts_found = [line.strip() for line in result.stdout.split('\n') if line.strip()]
                        logger.info(f"fc-list 找到 {len(fonts_found)} 个中文字体")
                        for font_file in fonts_found[:10]:
                            candidates.append((font_file, FONT_NAME, [None, 0]))
                except Exception as e:
                    logger.warning(f"fc-list 执行失败: {e}")
                
                hardcoded_linux_fonts = [
                    ('/usr/share/fonts/truetype/wqy/wqy-microhei.ttc', [None, 0]),
                    ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', [None, 0]),
                    ('/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc', [0]),
                    ('/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf', [None]),
                    ('/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf', [None]),
                    ('/usr/share/fonts/truetype/arphic/uming.ttc', [0]),
                    ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf', [None]),
                ]
                
                for path, indices in hardcoded_linux_fonts:
                    candidates.append((path, FONT_NAME, indices))
                    
            elif system == 'Darwin':
                candidates = [
                    ('/System/Library/Fonts/PingFang.ttc', None),
                    ('/System/Library/Fonts/STHeiti Light.ttc', None),
                    ('/System/Library/Fonts/Hiragino Sans GB.ttc', None),
                    ('/Library/Arial Unicode.ttf', None),
                ]
                for i, (path, _) in enumerate(candidates):
                    candidates[i] = (path, FONT_NAME, [None])
            
            seen_paths = set()
            for font_path, name, indices in candidates:
                if font_path in seen_paths:
                    continue
                seen_paths.add(font_path)
                
                for idx in indices:
                    success, msg = try_register(font_path, name, idx)
                    logger.info(f"尝试 {font_path} (index={idx}): {msg}")
                    if success:
                        logger.info(f"✅ 中文字体注册成功: {font_path}")
                        return True
            
            logger.error("❌ 所有中文字体注册尝试均失败，将使用默认字体")
            return False
        
        font_registered = register_chinese_font()
        
        output_path = os.path.join(OUTPUT_DIR, filename)
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )
        
        styles = getSampleStyleSheet()
        
        acunetix_dark = colors.HexColor('#1a2332')
        acunetix_primary = colors.HexColor('#ff6b35')
        acunetix_blue = colors.HexColor('#3b82f6')
        acunetix_bg = colors.HexColor('#f8fafc')
        acunetix_border = colors.HexColor('#e2e8f0')
        acunetix_text = colors.HexColor('#334155')
        acunetix_text_light = colors.HexColor('#64748b')
        
        critical_color = colors.HexColor('#dc2626')
        high_color = colors.HexColor('#ea580c')
        medium_color = colors.HexColor('#d97706')
        low_color = colors.HexColor('#16a34a')
        info_color = colors.HexColor('#0891b2')
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            spaceAfter=6,
            alignment=0,
            textColor=colors.white,
            fontName=FONT_NAME if font_registered else 'Helvetica-Bold',
            leading=28
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#94a3b8'),
            fontName=FONT_NAME if font_registered else 'Helvetica',
            spaceAfter=0
        )
        
        section_title_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=16,
            spaceAfter=10,
            textColor=acunetix_dark,
            fontName=FONT_NAME if font_registered else 'Helvetica-Bold',
            borderPadding=(0, 0, 4, 0)
        )
        
        heading3_style = ParagraphStyle(
            'Heading3',
            parent=styles['Heading3'],
            fontSize=12,
            spaceBefore=8,
            spaceAfter=6,
            textColor=acunetix_text,
            fontName=FONT_NAME if font_registered else 'Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'NormalCN',
            parent=styles['Normal'],
            fontSize=9.5,
            fontName=FONT_NAME if font_registered else 'Helvetica',
            textColor=acunetix_text,
            leading=14
        )
        
        small_style = ParagraphStyle(
            'SmallCN',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName=FONT_NAME if font_registered else 'Helvetica',
            textColor=acunetix_text_light,
            leading=12
        )
        
        mono_style = ParagraphStyle(
            'Mono',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Courier',
            textColor=acunetix_text,
            backColor=colors.HexColor('#f1f5f9'),
            leading=11
        )
        
        vuln_title_style = ParagraphStyle(
            'VulnTitle',
            parent=styles['Heading3'],
            fontSize=11,
            spaceBefore=10,
            spaceAfter=6,
            fontName=FONT_NAME if font_registered else 'Helvetica-Bold',
            textColor=acunetix_dark,
            leftIndent=0
        )
        
        summary = self._get_summary(task)
        
        story = []
        
        page_width = A4[0] - 40*mm
        
        header_data = [[
            Paragraph("<b>🛡️ Web应用程序漏洞检测报告</b>", title_style),
        ], [
            Paragraph("Aegis Security Scanner | 自动化漏洞检测与攻击验证系统", subtitle_style),
        ]]
        header_table = Table(header_data, colWidths=[page_width])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), acunetix_dark),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (0, 0), 24),
            ('BOTTOMPADDING', (0, 0), (0, 0), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 16),
            ('LEFTPADDING', (0, 0), (-1, -1), 24),
            ('RIGHTPADDING', (0, 0), (-1, -1), 24),
        ]))
        story.append(header_table)
        
        scan_duration = ""
        if task.created_at and task.updated_at:
            duration = task.updated_at - task.created_at
            total_seconds = int(duration.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            if hours > 0:
                scan_duration = f"{hours}小时{minutes}分钟{seconds}秒"
            elif minutes > 0:
                scan_duration = f"{minutes}分钟{seconds}秒"
            else:
                scan_duration = f"{seconds}秒"
        
        status_display = self._get_status_display(task.status)
        status_color = colors.HexColor('#16a34a') if task.status == 'COMPLETED' else colors.HexColor('#64748b')
        
        basic_info_data = [
            [Paragraph("<b>扫描目标</b>", small_style), Paragraph(task.target_url or 'N/A', normal_style)],
            [Paragraph("<b>任务ID</b>", small_style), Paragraph(str(task.id), normal_style)],
            [Paragraph("<b>开始时间</b>", small_style), Paragraph(task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else "N/A", normal_style)],
            [Paragraph("<b>结束时间</b>", small_style), Paragraph(task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else "N/A", normal_style)],
            [Paragraph("<b>扫描耗时</b>", small_style), Paragraph(scan_duration or "N/A", normal_style)],
            [Paragraph("<b>任务状态</b>", small_style), Paragraph(f"<b>{status_display}</b>", ParagraphStyle('Status', parent=normal_style, textColor=status_color))],
        ]
        
        col_w = [page_width * 0.25, page_width * 0.75]
        basic_table = Table(basic_info_data, colWidths=col_w, rowHeights=[22]*len(basic_info_data))
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), acunetix_text),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, acunetix_border),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('ROUNDEDCORNERS', [3, 3, 3, 3]),
        ]))
        story.append(Spacer(1, 16))
        story.append(basic_table)
        
        risk_score = (
            summary['critical'] * 10 +
            summary['high'] * 7 +
            summary['medium'] * 4 +
            summary['low'] * 1
        )
        
        if risk_score >= 70:
            risk_label = "高风险"
            risk_bar_color = critical_color
        elif risk_score >= 40:
            risk_label = "中风险"
            risk_bar_color = high_color
        elif risk_score > 0:
            risk_label = "低风险"
            risk_bar_color = medium_color
        else:
            risk_label = "无风险"
            risk_bar_color = low_color
        
        severity_configs = [
            ('严重', 'CRITICAL', summary['critical'], critical_color),
            ('高危', 'HIGH', summary['high'], high_color),
            ('中危', 'MEDIUM', summary['medium'], medium_color),
            ('低危', 'LOW', summary['low'], low_color),
            ('信息', 'INFO', summary['info'], info_color),
        ]
        
        stats_header = [
            Paragraph("<b>风险等级</b>", ParagraphStyle('StatsHeader', parent=small_style, textColor=colors.white)),
            Paragraph("<b>数量</b>", ParagraphStyle('StatsHeader', parent=small_style, textColor=colors.white)),
            Paragraph("<b>占比</b>", ParagraphStyle('StatsHeader', parent=small_style, textColor=colors.white)),
        ]
        
        stats_rows = [stats_header]
        total_vulns = max(summary['total'], 1)
        for label, eng_label, count, color in severity_configs:
            pct = (count / total_vulns) * 100 if total_vulns > 0 else 0
            count_cell = Paragraph(f"<b>{count}</b>", ParagraphStyle('CountCell', parent=normal_style, textColor=color))
            pct_cell = Paragraph(f"{pct:.1f}%", small_style)
            label_cell = Paragraph(f"<b>{label}</b><br/><font size=7 color='#94a3b8'>{eng_label}</font>", 
                                   ParagraphStyle('LabelCell', parent=normal_style, fontSize=9))
            stats_rows.append([label_cell, count_cell, pct_cell])
        
        stats_rows.append([
            Paragraph("<b>总计</b>", ParagraphStyle('TotalLabel', parent=normal_style, fontSize=10)),
            Paragraph(f"<b>{summary['total']}</b>", ParagraphStyle('TotalCount', parent=normal_style, fontSize=10, textColor=acunetix_dark)),
            Paragraph("<b>100%</b>", ParagraphStyle('TotalPct', parent=normal_style, fontSize=10)),
        ])
        
        stats_col_w = [page_width * 0.35, page_width * 0.25, page_width * 0.25]
        stats_table = Table(stats_rows, colWidths=stats_col_w)
        stats_style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), acunetix_dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, acunetix_border),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
            ('LINEBELOW', (0, -2), (-1, -2), 1.5, acunetix_dark),
        ]
        
        for i, (_, _, count, color) in enumerate(severity_configs, start=1):
            if count > 0:
                stats_style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fff7ed')))
        
        stats_table.setStyle(TableStyle(stats_style_commands))
        
        story.append(Spacer(1, 16))
        story.append(Paragraph("📊 漏洞统计摘要", section_title_style))
        story.append(stats_table)
        
        risk_summary_data = [[
            Paragraph("<b>综合风险评分</b>", normal_style),
            Paragraph(f"<b><font size=16 color='{risk_bar_color.hexval()}'>{risk_score}</font></b><br/><font size=8 color='#64748b'>/ 100 分</font>", 
                     ParagraphStyle('RiskScore', parent=normal_style, alignment=1)),
            Paragraph(f"<b>{risk_label}</b>", ParagraphStyle('RiskLabel', parent=normal_style, textColor=risk_bar_color, alignment=1)),
        ]]
        risk_table = Table(risk_summary_data, colWidths=[page_width*0.4, page_width*0.3, page_width*0.3])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fefce8')),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#eab308')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(Spacer(1, 10))
        story.append(risk_table)
        
        if task.vulnerabilities:
            story.append(Spacer(1, 16))
            story.append(Paragraph("🔍 漏洞详情", section_title_style))
            
            for idx, vuln in enumerate(task.vulnerabilities, start=1):
                severity_display = self._get_severity_display(vuln.severity)
                sev_lower = (vuln.severity or 'info').lower()
                
                sev_color_map = {
                    'critical': critical_color,
                    'high': high_color,
                    'medium': medium_color,
                    'low': low_color,
                    'info': info_color,
                }
                sev_bg_map = {
                    'critical': colors.HexColor('#fef2f2'),
                    'high': colors.HexColor('#fff7ed'),
                    'medium': colors.HexColor('#fffbeb'),
                    'low': colors.HexColor('#f0fdf4'),
                    'info': colors.HexColor('#f0f9ff'),
                }
                
                current_sev_color = sev_color_map.get(sev_lower, info_color)
                current_sev_bg = sev_bg_map.get(sev_lower, colors.HexColor('#f0f9ff'))
                
                vuln_header_data = [[
                    Paragraph(f"<font color='{current_sev_color.hexval()}'>●</font>  <b>{idx}. [{severity_display}] {vuln.vuln_name or vuln.vuln_type or '未知漏洞'}</b>", vuln_title_style),
                ]]
                vuln_header_table = Table(vuln_header_data, colWidths=[page_width])
                vuln_header_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), current_sev_bg),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                    ('LEFTPADDING', (0, 0), (-1, -1), 14),
                    ('BOX', (0, 0), (-1, -1), 0, colors.white),
                    ('LINEBELOW', (0, 0), (-1, -1), 2, current_sev_color),
                ]))
                story.append(vuln_header_table)
                
                cvss_display = "N/A"
                if vuln.cvss_score is not None:
                    cvss_display = f"{vuln.cvss_score}/10"
                
                payload_display = "N/A"
                if vuln.payload:
                    payload_str = str(vuln.payload)
                    if len(payload_str) > 300:
                        payload_display = payload_str[:300] + "..."
                    else:
                        payload_display = payload_str
                
                vuln_detail_data = [
                    [Paragraph("<b>漏洞类型</b>", small_style), Paragraph(vuln.vuln_type or 'N/A', normal_style)],
                    [Paragraph("<b>风险等级</b>", small_style), 
                     Paragraph(f"<font color='{current_sev_color.hexval()}'><b>{severity_display.upper()}</b></font>  ({(vuln.severity or 'N/A').upper()})", normal_style)],
                    [Paragraph("<b>CVSS评分</b>", small_style), Paragraph(cvss_display, normal_style)],
                    [Paragraph("<b>触发URL</b>", small_style), Paragraph(f"<font face='Courier' size=8>{vuln.url or 'N/A'}</font>", normal_style)],
                    [Paragraph("<b>HTTP方法</b>", small_style), Paragraph(f"<b>{vuln.method or 'N/A'}</b>", normal_style)],
                    [Paragraph("<b>注入参数</b>", small_style), Paragraph(f"<font face='Courier' size=8>{vuln.parameter or 'N/A'}</font>", normal_style)],
                ]
                
                if vuln.detected_at:
                    vuln_detail_data.append([
                        Paragraph("<b>检测时间</b>", small_style), 
                        Paragraph(vuln.detected_at.strftime("%Y-%m-%d %H:%M:%S"), normal_style)
                    ])
                
                vuln_detail_data.append([Paragraph("<b>攻击载荷</b>", small_style), 
                                         Paragraph(f"<font face='Courier' size=7.5 color='#475569'>{payload_display}</font>", 
                                                  ParagraphStyle('PayloadStyle', parent=normal_style, leading=11))])
                
                if vuln.description:
                    vuln_detail_data.append([Paragraph("<b>漏洞描述</b>", small_style), 
                                             Paragraph(vuln.description, normal_style)])
                
                if vuln.remediation:
                    vuln_detail_data.append([Paragraph("<b>修复建议</b>", small_style), 
                                             Paragraph(f"<font color='#16a34a'>{vuln.remediation}</font>", normal_style)])
                
                vuln_detail_table = Table(vuln_detail_data, colWidths=[page_width * 0.22, page_width * 0.78])
                vuln_detail_style_cmds = [
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fafafa')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), acunetix_text),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ]
                
                vuln_detail_table.setStyle(TableStyle(vuln_detail_style_cmds))
                story.append(vuln_detail_table)
                story.append(Spacer(1, 10))
        else:
            story.append(Spacer(1, 16))
            no_vuln_data = [[Paragraph("<font size=14 color='#16a34a'>✅ </font><b>未发现任何漏洞</b>", 
                                        ParagraphStyle('NoVuln', parent=normal_style, alignment=1, fontSize=13))],
                           [Paragraph("目标应用程序在本次扫描中未检测到安全漏洞。", 
                                      ParagraphStyle('NoVulnDesc', parent=small_style, alignment=1))]]
            no_vuln_table = Table(no_vuln_data, colWidths=[page_width])
            no_vuln_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0fdf4')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#bbf7d0')),
                ('TOPPADDING', (0, 0), (-1, -1), 20),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(no_vuln_table)
        
        story.append(Spacer(1, 24))
        story.append(HRFlowable(width="100%", thickness=1, color=acunetix_border, spaceBefore=5, spaceAfter=12))
        
        footer_data = [[
            Paragraph(f"<font color='#94a3b8' size=8>本报告由 Aegis Web应用程序漏洞检测系统自动生成 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>", 
                     ParagraphStyle('Footer', parent=small_style, alignment=1)),
        ]]
        footer_table = Table(footer_data, colWidths=[page_width])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(footer_table)
        
        doc.build(story)
        
        return output_path

    def generate(self, task: ScanTask, format: str = "html") -> tuple[str, str]:
        """
        根据指定格式生成报告
        
        Args:
            task: 扫描任务对象
            format: 导出格式 (html/pdf/markdown/excel/json)
            
        Returns:
            (文件路径, 文件名) 元组
        """
        format = format.lower()
        
        # 文件扩展名映射
        extensions = {
            "html": ".html",
            "pdf": ".pdf",
            "markdown": ".md",
            "excel": ".xlsx",
            "json": ".json"
        }
        
        # MIME 类型映射
        mime_types = {
            "html": "text/html",
            "pdf": "application/pdf",
            "markdown": "text/markdown",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json"
        }
        
        if format not in extensions:
            format = "html"
        
        filename = f"report_{task.id}{extensions[format]}"
        
        # 根据格式调用相应的生成方法
        generators = {
            "html": self.generate_html,
            "markdown": self.generate_markdown,
            "json": self.generate_json,
            "excel": self.generate_excel,
            "pdf": self.generate_pdf
        }
        
        file_path = generators[format](task, filename)
        
        return file_path, filename