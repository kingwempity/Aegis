"""
扫描任务管理模块视图
提供扫描任务的创建、查询、取消和管理功能
"""
import re
from django.utils import timezone
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import ScanTask, Vulnerability, ScanResult
from .serializers import ScanTaskSerializer, VulnerabilitySerializer
try:
    from vuln_scanner.scanner.executor import get_executor
except ImportError:
    # 如果导入失败，提供一个简化的实现
    def get_executor():
        return None


class ScanTaskCreateView(APIView):
    """
    扫描任务创建视图
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        创建新的扫描任务

        请求参数:
        - target_url: 目标URL
        - task_name: 任务名称 (可选)
        - scan_profile: 扫描配置模板
        - custom_modules: 自定义检测模块
        - auth_token: 认证Token
        - auth_cookies: 认证Cookie
        - max_depth: 爬虫最大深度
        - max_pages: 最大扫描页面数
        - timeout: 请求超时时间
        - user_agent: 自定义User-Agent
        - headers: 自定义请求头
        - exclude_patterns: 排除的URL模式

        返回:
        - 成功: 任务ID和状态
        - 失败: 错误信息
        """
        serializer = ScanTaskSerializer(data=request.data)
        if serializer.is_valid():
            # 设置创建者
            scan_task = serializer.save(created_by=request.user)

            return Response({
                'code': 200,
                'message': 'Task created successfully',
                'data': {
                    'task_id': scan_task.task_id,
                    'status': scan_task.status,
                    'created_at': scan_task.created_at.isoformat()
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'code': 400,
                'message': 'Invalid request data',
                'data': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)


class ScanTaskStatusView(APIView):
    """
    扫描任务状态查询视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        """
        获取指定任务的状态和进度

        路径参数:
        - task_id: 任务ID

        返回:
        - 任务状态、进度、当前阶段等信息
        """
        try:
            scan_task = get_object_or_404(
                ScanTask,
                task_id=task_id,
                created_by=request.user
            )

            # 计算预计完成时间（简化版）
            estimated_completion = None
            if scan_task.started_at and scan_task.status == 'running':
                duration_so_far = scan_task.duration_seconds()
                if duration_so_far > 0 and scan_task.progress > 0:
                    total_estimated = (duration_so_far * 100) / scan_task.progress
                    remaining = total_estimated - duration_so_far
                    estimated_completion = scan_task.started_at + timezone.timedelta(seconds=total_estimated)

            data = {
                'task_id': scan_task.task_id,
                'status': scan_task.status,
                'progress': scan_task.progress,
                'current_phase': self._get_current_phase(scan_task),
                'started_at': scan_task.started_at.isoformat() if scan_task.started_at else None,
                'estimated_completion': estimated_completion.isoformat() if estimated_completion else None
            }

            return Response({
                'code': 200,
                'message': 'Success',
                'data': data
            }, status=status.HTTP_200_OK)

        except ScanTask.DoesNotExist:
            return Response({
                'code': 404,
                'message': 'Task not found',
                'data': {}
            }, status=status.HTTP_404_NOT_FOUND)

    def _get_current_phase(self, scan_task):
        """
        根据任务进度和状态确定当前阶段
        """
        if scan_task.status != 'running':
            return 'N/A'

        progress = scan_task.progress
        if progress < 20:
            return 'Initializing'
        elif progress < 40:
            return 'Crawling'
        elif progress < 60:
            return 'SQL Injection Testing'
        elif progress < 80:
            return 'XSS Testing'
        elif progress < 100:
            return 'Report Generation'
        else:
            return 'Completed'


class ScanTaskListView(APIView):
    """
    扫描任务列表视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        获取当前用户的所有扫描任务列表

        查询参数:
        - page: 页码
        - page_size: 每页数量
        - status_filter: 状态过滤
        - sort_by: 排序字段
        - order: 排序方向

        返回:
        - 分页的任务列表
        """
        # 获取查询参数
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        status_filter = request.query_params.get('status_filter', 'all')
        sort_by = request.query_params.get('sort_by', 'created_at')
        order = request.query_params.get('order', 'desc')

        # 验证参数
        if page < 1:
            page = 1
        if page_size < 1 or page_size > 50:
            page_size = 10

        # 构建查询
        queryset = ScanTask.objects.filter(created_by=request.user)

        # 状态过滤
        if status_filter != 'all':
            queryset = queryset.filter(status=status_filter)

        # 排序
        if sort_by not in ['created_at', 'completed_at', 'status']:
            sort_by = 'created_at'

        order_by = f"{'-' if order == 'desc' else ''}{sort_by}"
        queryset = queryset.order_by(order_by)

        # 分页
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        tasks = queryset[start:end]

        # 序列化数据
        task_data = []
        for task in tasks:
            task_data.append({
                'task_id': task.task_id,
                'task_name': task.task_name,
                'target_url': task.target_url,
                'status': task.status,
                'progress': task.progress,
                'vulnerabilities_found': task.vulnerabilities_found,
                'created_at': task.created_at.isoformat(),
                'completed_at': task.completed_at.isoformat() if task.completed_at else None
            })

        return Response({
            'code': 200,
            'message': 'Success',
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'tasks': task_data
            }
        }, status=status.HTTP_200_OK)


class ScanTaskCancelView(APIView):
    """
    扫描任务取消视图
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, task_id):
        """
        取消指定的扫描任务

        路径参数:
        - task_id: 任务ID

        返回:
        - 取消结果
        """
        try:
            scan_task = get_object_or_404(
                ScanTask,
                task_id=task_id,
                created_by=request.user
            )

            if scan_task.status == 'completed':
                return Response({
                    'code': 400,
                    'message': 'Task already completed',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

            if not scan_task.can_cancel():
                return Response({
                    'code': 400,
                    'message': 'Task cannot be cancelled',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

            # 通过执行器取消任务
            executor = get_executor()
            success, message = executor.cancel_task(scan_task.task_id, request.user)

            if success:
                return Response({
                    'code': 200,
                    'message': 'Task cancelled successfully',
                    'data': {
                        'task_id': scan_task.task_id,
                        'status': 'cancelled'
                    }
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'code': 400,
                    'message': message,
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

        except ScanTask.DoesNotExist:
            return Response({
                'code': 404,
                'message': 'Task not found',
                'data': {}
            }, status=status.HTTP_404_NOT_FOUND)


class ScanTaskDetailView(APIView):
    """
    扫描任务详情视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        """
        获取指定任务的完整详细信息

        路径参数:
        - task_id: 任务ID

        返回:
        - 任务的完整详情
        """
        try:
            scan_task = get_object_or_404(
                ScanTask,
                task_id=task_id,
                created_by=request.user
            )

            serializer = ScanTaskSerializer(scan_task)

            return Response({
                'code': 200,
                'message': 'Success',
                'data': serializer.data
            }, status=status.HTTP_200_OK)

        except ScanTask.DoesNotExist:
            return Response({
                'code': 404,
                'message': 'Task not found',
                'data': {}
            }, status=status.HTTP_404_NOT_FOUND)


class ScanReportView(APIView):
    """
    扫描报告视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        """
        获取扫描报告

        路径参数:
        - task_id: 任务ID

        查询参数:
        - format: 报告格式 (json/pdf)

        返回:
        - 扫描报告数据或PDF文件
        """
        try:
            scan_task = get_object_or_404(
                ScanTask,
                task_id=task_id,
                created_by=request.user
            )

            if scan_task.status != 'completed':
                return Response({
                    'code': 400,
                    'message': 'Task not completed yet',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

            # 获取报告格式
            report_format = request.query_params.get('format', 'json')

            if report_format == 'pdf':
                # TODO: 生成PDF报告
                return Response({
                    'code': 501,
                    'message': 'PDF report generation not implemented yet',
                    'data': {}
                }, status=status.HTTP_501_NOT_IMPLEMENTED)

            # 生成JSON报告
            vulnerabilities = Vulnerability.objects.filter(task=scan_task)
            vuln_serializer = VulnerabilitySerializer(vulnerabilities, many=True)

            # 统计漏洞数量
            summary = {
                'total_vulnerabilities': vulnerabilities.count(),
                'critical': vulnerabilities.filter(risk_level='critical').count(),
                'high': vulnerabilities.filter(risk_level='high').count(),
                'medium': vulnerabilities.filter(risk_level='medium').count(),
                'low': vulnerabilities.filter(risk_level='low').count(),
                'pages_scanned': scan_task.pages_scanned,
                'modules_executed': len(scan_task.custom_modules) if scan_task.custom_modules else 5
            }

            # 获取技术栈信息
            technology_stack = {}
            try:
                scan_result = ScanResult.objects.get(task=scan_task)
                technology_stack = scan_result.technology_stack or {}
            except ScanResult.DoesNotExist:
                pass

            report_data = {
                'task_id': scan_task.task_id,
                'target_url': scan_task.target_url,
                'scan_time': scan_task.completed_at.isoformat() if scan_task.completed_at else timezone.now().isoformat(),
                'scan_duration': scan_task.duration_seconds(),
                'summary': summary,
                'vulnerabilities': vuln_serializer.data,
                'technology_stack': technology_stack
            }

            return Response({
                'code': 200,
                'message': 'Report generated',
                'data': report_data
            }, status=status.HTTP_200_OK)

        except ScanTask.DoesNotExist:
            return Response({
                'code': 404,
                'message': 'Task not found',
                'data': {}
            }, status=status.HTTP_404_NOT_FOUND)


class ScanReportExportView(APIView):
    """
    扫描报告导出视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        """
        导出扫描报告

        路径参数:
        - task_id: 任务ID

        查询参数:
        - format: 导出格式 (pdf/excel/html/markdown)
        - include_evidence: 是否包含证据
        - include_screenshots: 是否包含截图

        返回:
        - 导出的报告文件
        """
        import logging
        logger = logging.getLogger(__name__)

        logger.info(f"=== 开始处理导出请求 ===")
        logger.info(f"task_id: {task_id}")
        logger.info(f"request.user: {request.user}")
        logger.info(f"request.user.is_authenticated: {request.user.is_authenticated}")
        logger.info(f"request.method: {request.method}")
        logger.info(f"request.GET: {dict(request.GET)}")
        logger.info(f"request.path: {request.path}")
        logger.info(f"request.META.get('HTTP_AUTHORIZATION'): {request.META.get('HTTP_AUTHORIZATION', 'No auth header')}")

        try:
            # 记录请求信息以便调试
            logger.info(f"导出报告请求: task_id={task_id}, user={request.user.username if request.user.is_authenticated else 'Anonymous'}")
            
            # 先检查任务是否存在（不检查用户）
            task_exists = ScanTask.objects.filter(task_id=task_id).exists()
            if not task_exists:
                logger.warning(f"任务不存在: task_id={task_id}")
                return Response({
                    'code': 404,
                    'message': f'任务不存在 (任务ID: {task_id})',
                    'data': {}
                }, status=status.HTTP_404_NOT_FOUND)
            
            # 检查任务是否属于当前用户
            try:
                scan_task = ScanTask.objects.get(
                    task_id=task_id,
                    created_by=request.user
                )
                logger.info(f"找到任务: {scan_task.task_id}, 状态: {scan_task.status}")
            except ScanTask.DoesNotExist:
                # 任务存在但不属于当前用户
                task_owner = ScanTask.objects.filter(task_id=task_id).first()
                owner_name = task_owner.created_by.username if task_owner else "未知"
                logger.warning(f"用户无权限访问任务: task_id={task_id}, user={request.user.username}, owner={owner_name}")
                return Response({
                    'code': 403,
                    'message': f'您没有权限访问该任务 (任务ID: {task_id})',
                    'data': {}
                }, status=status.HTTP_403_FORBIDDEN)

            if scan_task.status != 'completed':
                return Response({
                    'code': 400,
                    'message': 'Task not completed yet',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

            # 获取导出参数
            export_format = request.GET.get('format', 'pdf')
            include_evidence = request.GET.get('include_evidence', 'false').lower() == 'true'
            include_screenshots = request.GET.get('include_screenshots', 'true').lower() == 'true'

            # 验证导出格式
            supported_formats = ['pdf', 'excel', 'html', 'markdown', 'json']
            if export_format not in supported_formats:
                return Response({
                    'code': 400,
                    'message': f'Unsupported export format. Supported formats: {", ".join(supported_formats)}',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

            # 实现导出逻辑
            if export_format == 'html':
                return self._export_html_report(scan_task, include_evidence, include_screenshots)
            elif export_format == 'pdf':
                return self._export_pdf_report(scan_task, include_evidence, include_screenshots)
            elif export_format == 'excel':
                return self._export_excel_report(scan_task, include_evidence, include_screenshots)
            elif export_format == 'markdown':
                return self._export_markdown_report(scan_task, include_evidence, include_screenshots)
            elif export_format == 'json':
                return self._export_json_report(scan_task, include_evidence, include_screenshots)
            else:
                return Response({
                    'code': 400,
                    'message': 'Unsupported export format',
                    'data': {}
                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # 记录异常以便调试
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"导出报告时发生错误: {str(e)}")
            return Response({
                'code': 500,
                'message': f'导出报告失败: {str(e)}',
                'data': {}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _export_html_report(self, scan_task, include_evidence, include_screenshots):
        """导出HTML报告"""
        from django.http import HttpResponse

        vulnerabilities = Vulnerability.objects.filter(task=scan_task)

        # 统计数据
        summary = {
            "total_vulnerabilities": vulnerabilities.count(),
            "critical": vulnerabilities.filter(risk_level="critical").count(),
            "high": vulnerabilities.filter(risk_level="high").count(),
            "medium": vulnerabilities.filter(risk_level="medium").count(),
            "low": vulnerabilities.filter(risk_level="low").count(),
            "pages_scanned": scan_task.pages_scanned,
            "modules_executed": len(scan_task.custom_modules) if scan_task.custom_modules else 5
        }

        # 生成HTML内容
        html_content = f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>漏洞扫描报告 - {scan_task.task_id}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background: #f8f9fa; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
                .summary {{ background: #e9ecef; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .vulnerability {{ border: 1px solid #dee2e6; padding: 15px; margin-bottom: 10px; border-radius: 5px; }}
                .critical {{ border-left: 5px solid #dc3545; }}
                .high {{ border-left: 5px solid #fd7e14; }}
                .medium {{ border-left: 5px solid #ffc107; }}
                .low {{ border-left: 5px solid #28a745; }}
                .info {{ border-left: 5px solid #17a2b8; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #dee2e6; padding: 8px; text-align: left; }}
                th {{ background: #f8f9fa; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Web应用程序漏洞检测报告</h1>
                <p><strong>任务ID:</strong> {scan_task.task_id}</p>
                <p><strong>目标URL:</strong> {scan_task.target_url}</p>
                <p><strong>扫描时间:</strong> {scan_task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan_task.completed_at else "N/A"}</p>
                <p><strong>扫描持续时间:</strong> {scan_task.duration_seconds()} 秒</p>
            </div>

            <div class="summary">
                <h2>扫描摘要</h2>
                <table>
                    <tr><th>总漏洞数</th><td>{summary["total_vulnerabilities"]}</td></tr>
                    <tr><th>危急漏洞</th><td>{summary["critical"]}</td></tr>
                    <tr><th>高危漏洞</th><td>{summary["high"]}</td></tr>
                    <tr><th>中危漏洞</th><td>{summary["medium"]}</td></tr>
                    <tr><th>低危漏洞</th><td>{summary["low"]}</td></tr>
                    <tr><th>扫描页面数</th><td>{summary["pages_scanned"]}</td></tr>
                    <tr><th>执行模块数</th><td>{summary["modules_executed"]}</td></tr>
                </table>
            </div>

            <h2>漏洞详情</h2>
        """

        for vuln in vulnerabilities:
            risk_class = vuln.risk_level.lower()
            html_content += f"""
            <div class="vulnerability {risk_class}">
                <h3>{vuln.name} ({vuln.risk_level.upper()})</h3>
                <p><strong>类型:</strong> {vuln.type}</p>
                <p><strong>URL:</strong> {vuln.url}</p>
                <p><strong>参数:</strong> {vuln.parameter or "N/A"}</p>
                <p><strong>CVSS评分:</strong> {vuln.cvss_score}</p>
                <p><strong>证据:</strong> {vuln.evidence}</p>
                <p><strong>修复建议:</strong> {vuln.remediation}</p>
        """

            # 如果需要包含截图信息
            if include_screenshots and getattr(vuln, "screenshots", None):
                html_content += "<p><strong>截图/证据链接:</strong></p><ul>"
                for shot in vuln.screenshots:
                    desc = shot.get("description") or "截图"
                    url = shot.get("url") or "#"
                    html_content += f'<li><a href="{url}">{desc}</a></li>'
                html_content += "</ul>"

            html_content += """
            </div>
            """

        html_content += """
        </body>
        </html>
        """

        response = HttpResponse(html_content, content_type="text/html; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="scan_report_{scan_task.task_id}.html"'
        return response

    def _export_markdown_report(self, scan_task, include_evidence, include_screenshots):
        """导出Markdown报告（与功能需求文档中的报告结构对齐）"""
        from django.http import HttpResponse

        vulnerabilities = Vulnerability.objects.filter(task=scan_task)

        # 统计数据
        summary = {
            "total_vulnerabilities": vulnerabilities.count(),
            "critical": vulnerabilities.filter(risk_level="critical").count(),
            "high": vulnerabilities.filter(risk_level="high").count(),
            "medium": vulnerabilities.filter(risk_level="medium").count(),
            "low": vulnerabilities.filter(risk_level="low").count(),
            "pages_scanned": getattr(scan_task, "pages_scanned", None),
            "modules_executed": len(scan_task.custom_modules) if scan_task.custom_modules else 5
        }

        lines = []
        lines.append("# 漏洞检测报告")
        lines.append("")
        lines.append(f"- 任务ID: {scan_task.task_id}")
        lines.append(f"- 目标URL: {scan_task.target_url}")
        if scan_task.completed_at:
            lines.append(f"- 扫描时间: {scan_task.completed_at.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"- 扫描持续时间: {scan_task.duration_seconds()} 秒")
        lines.append("")

        lines.append("## 扫描概览")
        lines.append(f"- 扫描页面数: {summary['pages_scanned']}")
        lines.append(f"- 执行模块数: {summary['modules_executed']}")
        lines.append("")

        lines.append("## 漏洞统计")
        lines.append(f"- 总漏洞数: {summary['total_vulnerabilities']}")
        lines.append(f"- 危急: {summary['critical']}")
        lines.append(f"- 高危: {summary['high']}")
        lines.append(f"- 中危: {summary['medium']}")
        lines.append(f"- 低危: {summary['low']}")
        lines.append("")

        lines.append("## 漏洞详情")
        if not vulnerabilities:
            lines.append("")
            lines.append("当前报告未发现漏洞。")
        else:
            for idx, vuln in enumerate(vulnerabilities, start=1):
                lines.append("")
                lines.append(f"### {idx}. [{vuln.risk_level.upper()}] {vuln.name} (CVSS: {vuln.cvss_score})")
                lines.append(f"- 漏洞类型: {vuln.type}")
                lines.append(f"- URL: {vuln.url}")
                lines.append(f"- 请求方法: {vuln.method}")
                if vuln.parameter:
                    lines.append(f"- 参数: {vuln.parameter}")

                lines.append("")
                lines.append("**漏洞描述**")
                lines.append("")
                lines.append(vuln.description or "")
                lines.append("")

                lines.append("**修复建议**")
                lines.append("")
                lines.append(vuln.remediation or "")

                if include_evidence:
                    lines.append("")
                    lines.append("**攻击证据**")
                    lines.append("")
                    if vuln.payload:
                        lines.append("攻击载荷:")
                        lines.append("")
                        lines.append("```")
                        lines.append(vuln.payload)
                        lines.append("```")
                        lines.append("")
                    lines.append(vuln.evidence or "（无证据详情）")

                # 攻击步骤（如果有）
                attack_steps = getattr(vuln, "attack_steps", None)
                if include_evidence and attack_steps:
                    lines.append("")
                    lines.append("**攻击过程**")
                    lines.append("")
                    for step in attack_steps:
                        action = step.get("action")
                        code = step.get("response_code")
                        rt = step.get("response_time_ms")
                        lines.append(f"- {action} （响应码: {code}, 时间: {rt}ms）")

                # 截图 / 证据链接
                screenshots = getattr(vuln, "screenshots", None)
                if include_screenshots and screenshots:
                    lines.append("")
                    lines.append("**截图/证据链接**")
                    lines.append("")
                    for i, shot in enumerate(screenshots, start=1):
                        desc = shot.get("description") or "截图"
                        url = shot.get("url") or "#"
                        lines.append(f"- 图 {i}: {desc} ({url})")

                # 参考资料
                if vuln.references:
                    lines.append("")
                    lines.append("**参考资料**")
                    lines.append("")
                    for ref in vuln.references:
                        lines.append(f"- {ref}")

        lines.append("")
        lines.append("---")
        lines.append("报告由漏洞检测系统自动生成（Markdown 格式），可导入到 Word、Typora 等工具转换为 PDF、HTML 等格式。")

        markdown = "\n".join(lines)

        response = HttpResponse(markdown, content_type="text/markdown; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename=\"scan_report_{scan_task.task_id}.md\"'
        return response

    def _export_pdf_report(self, scan_task, include_evidence, include_screenshots):
        """导出PDF报告"""
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import os
        import platform
        import logging

        logger = logging.getLogger(__name__)

        # 注册中文字体
        font_name = 'Helvetica'  # 默认字体
        try:
            # 尝试注册系统字体
            font_registered = False
            if platform.system() == 'Windows':
                font_paths = [
                    r'C:\Windows\Fonts\simsun.ttc',
                    r'C:\Windows\Fonts\msyh.ttc',
                    r'C:\Windows\Fonts\simhei.ttf'
                ]
                for fp in font_paths:
                    if os.path.exists(fp):
                        pdfmetrics.registerFont(TTFont('SimSun', fp, subfontIndex=0 if fp.endswith('.ttc') else None))
                        font_name = 'SimSun'
                        font_registered = True
                        break
            elif platform.system() == 'Linux':
                font_paths = [
                    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
                    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
                    '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
                    '/usr/share/fonts/truetype/simhei.ttf'
                ]
                for fp in font_paths:
                    if os.path.exists(fp):
                        pdfmetrics.registerFont(TTFont('SimSun', fp))
                        font_name = 'SimSun'
                        font_registered = True
                        break
            elif platform.system() == 'Darwin':
                font_paths = [
                    '/System/Library/Fonts/PingFang.ttc',
                    '/Library/Fonts/Arial Unicode.ttf',
                    '/Library/Fonts/SimSun.ttf'
                ]
                for fp in font_paths:
                    if os.path.exists(fp):
                        pdfmetrics.registerFont(TTFont('SimSun', fp))
                        font_name = 'SimSun'
                        font_registered = True
                        break
            
            if not font_registered:
                logger.warning("未找到合适的中文字体，PDF可能无法正确显示中文")
        except Exception as e:
            logger.error(f"注册中文字体失败: {e}")

        try:
            # 创建PDF缓冲区
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                  rightMargin=50, leftMargin=50, 
                                  topMargin=50, bottomMargin=50)

            # 获取样式
            styles = getSampleStyleSheet()

            # 创建自定义样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=22,
                spaceAfter=30,
                alignment=1,
                textColor=colors.darkblue,
                fontName=font_name
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceBefore=15,
                spaceAfter=15,
                textColor=colors.darkblue,
                fontName=font_name,
                borderPadding=5,
                borderWidth=0,
                leftIndent=0
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                fontName=font_name,
                leading=14
            )

            table_header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontSize=10,
                fontName=font_name,
                textColor=colors.whitesmoke,
                alignment=1
            )

            # 获取数据
            vulnerabilities = Vulnerability.objects.filter(task=scan_task)

            # 统计数据
            summary = {
                "total_vulnerabilities": vulnerabilities.count(),
                "critical": vulnerabilities.filter(risk_level="critical").count(),
                "high": vulnerabilities.filter(risk_level="high").count(),
                "medium": vulnerabilities.filter(risk_level="medium").count(),
                "low": vulnerabilities.filter(risk_level="low").count(),
                "pages_scanned": getattr(scan_task, "pages_scanned", 0),
                "modules_executed": len(scan_task.custom_modules) if scan_task.custom_modules else 5
            }

            # 获取技术栈信息
            technology_stack = {}
            try:
                scan_result = ScanResult.objects.get(task=scan_task)
                technology_stack = scan_result.technology_stack or {}
            except ScanResult.DoesNotExist:
                pass

            # 构建PDF内容
            story = []

            # 标题
            story.append(Paragraph("Aegis 安全检测报告", title_style))
            story.append(Spacer(1, 20))

            # 基本信息
            story.append(Paragraph("1. 任务基本信息", heading_style))
            basic_info_data = [
                [Paragraph('<b>任务ID</b>', normal_style), Paragraph(scan_task.task_id, normal_style)],
                [Paragraph('<b>目标URL</b>', normal_style), Paragraph(scan_task.target_url, normal_style)],
                [Paragraph('<b>扫描时间</b>', normal_style), Paragraph(scan_task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan_task.completed_at else "N/A", normal_style)],
                [Paragraph('<b>扫描时长</b>', normal_style), Paragraph(f"{scan_task.duration_seconds()} 秒", normal_style)],
            ]

            basic_table = Table(basic_info_data, colWidths=[1.5*inch, 4.5*inch])
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            story.append(basic_table)
            story.append(Spacer(1, 20))

            # 扫描概览
            story.append(Paragraph("2. 扫描概览", heading_style))
            overview_data = [
                [Paragraph('<b>扫描页面数</b>', normal_style), Paragraph(str(summary['pages_scanned']), normal_style)],
                [Paragraph('<b>执行模块数</b>', normal_style), Paragraph(str(summary['modules_executed']), normal_style)],
                [Paragraph('<b>服务器信息</b>', normal_style), Paragraph(technology_stack.get('server', '未知'), normal_style)],
                [Paragraph('<b>编程语言</b>', normal_style), Paragraph(technology_stack.get('language', '未知'), normal_style)],
                [Paragraph('<b>框架/数据库</b>', normal_style), Paragraph(f"{technology_stack.get('framework', '未知')} / {technology_stack.get('database', '未知')}", normal_style)],
            ]

            overview_table = Table(overview_data, colWidths=[1.5*inch, 4.5*inch])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            story.append(overview_table)
            story.append(Spacer(1, 20))

            # 漏洞统计
            story.append(Paragraph("3. 漏洞统计", heading_style))
            stats_data = [
                [Paragraph('<b>风险等级</b>', table_header_style), Paragraph('<b>数量</b>', table_header_style)],
                [Paragraph('危急 (Critical)', normal_style), Paragraph(str(summary['critical']), normal_style)],
                [Paragraph('高危 (High)', normal_style), Paragraph(str(summary['high']), normal_style)],
                [Paragraph('中危 (Medium)', normal_style), Paragraph(str(summary['medium']), normal_style)],
                [Paragraph('低危 (Low)', normal_style), Paragraph(str(summary['low']), normal_style)],
                [Paragraph('<b>总计</b>', normal_style), Paragraph(f"<b>{summary['total_vulnerabilities']}</b>", normal_style)],
            ]

            stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.darkblue),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 20))

            # 漏洞详情
            if vulnerabilities.exists():
                story.append(PageBreak())
                story.append(Paragraph("4. 漏洞详情", heading_style))

                risk_colors = {
                    'critical': colors.red,
                    'high': colors.orange,
                    'medium': colors.goldenrod,
                    'low': colors.green,
                    'info': colors.blue
                }

                for idx, vuln in enumerate(vulnerabilities, start=1):
                    # 漏洞标题
                    vuln_title = f"{idx}. [{vuln.risk_level.upper()}] {vuln.name}"
                    story.append(Paragraph(vuln_title, ParagraphStyle(
                        'VulnTitle',
                        parent=styles['Heading3'],
                        fontSize=14,
                        textColor=risk_colors.get(vuln.risk_level, colors.black),
                        spaceBefore=10,
                        spaceAfter=10,
                        fontName=font_name
                    )))

                    # 漏洞详细信息表格
                    vuln_data = [
                        [Paragraph('<b>漏洞类型</b>', normal_style), Paragraph(vuln.type, normal_style)],
                        [Paragraph('<b>风险等级</b>', normal_style), Paragraph(vuln.risk_level.upper(), normal_style)],
                        [Paragraph('<b>CVSS评分</b>', normal_style), Paragraph(f"{vuln.cvss_score} ({vuln.cvss_vector or 'N/A'})", normal_style)],
                        [Paragraph('<b>发现URL</b>', normal_style), Paragraph(f"{vuln.method} {vuln.url}", normal_style)],
                    ]

                    if vuln.parameter:
                        vuln_data.append([Paragraph('<b>涉及参数</b>', normal_style), Paragraph(vuln.parameter, normal_style)])

                    if vuln.payload and include_evidence:
                        vuln_data.append([Paragraph('<b>攻击载荷</b>', normal_style), Paragraph(vuln.payload, normal_style)])

                    vuln_data.extend([
                        [Paragraph('<b>漏洞描述</b>', normal_style), Paragraph(vuln.description or 'N/A', normal_style)],
                        [Paragraph('<b>修复建议</b>', normal_style), Paragraph(vuln.remediation or 'N/A', normal_style)],
                        [Paragraph('<b>证据详情</b>', normal_style), Paragraph(vuln.evidence or '无证据详情', normal_style)],
                    ])

                    vuln_table = Table(vuln_data, colWidths=[1.2*inch, 4.8*inch])
                    vuln_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 8),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 5),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ]))
                    story.append(vuln_table)
                    story.append(Spacer(1, 15))

                    # 攻击步骤
                    if include_evidence and vuln.attack_steps:
                        story.append(Paragraph("<b>攻击过程:</b>", normal_style))
                        for step_idx, step in enumerate(vuln.attack_steps, start=1):
                            step_text = f"{step_idx}. {step.get('action', '')} (响应: {step.get('response_code', '')}, 耗时: {step.get('response_time_ms', '')}ms)"
                            story.append(Paragraph(step_text, ParagraphStyle('Step', parent=normal_style, leftIndent=20)))
                        story.append(Spacer(1, 10))

                    # 参考资料
                    if vuln.references:
                        story.append(Paragraph("<b>参考资料:</b>", normal_style))
                        for ref in vuln.references:
                            story.append(Paragraph(f"• {ref}", ParagraphStyle('Ref', parent=normal_style, leftIndent=20)))
                        story.append(Spacer(1, 10))
                    
                    # 漏洞之间添加分割线
                    if idx < vulnerabilities.count():
                        story.append(Spacer(1, 10))

            else:
                story.append(Paragraph("本次扫描未发现任何安全漏洞。", normal_style))

            # 生成PDF
            doc.build(story)

            # 创建HTTP响应
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="scan_report_{scan_task.task_id}.pdf"'
            response.write(buffer.getvalue())
            buffer.close()
            return response

        except Exception as e:
            logger.error(f"生成PDF报告失败: {e}", exc_info=True)
            from django.http import JsonResponse
            return JsonResponse({
                'code': 500,
                'message': f'Failed to generate PDF report: {str(e)}'
            }, status=500)

    def _export_excel_report(self, scan_task, include_evidence, include_screenshots):
        """
        导出Excel报告
        
        Args:
            scan_task: 扫描任务对象
            include_evidence: 是否包含攻击证据
            include_screenshots: 是否包含截图信息
            
        Returns:
            HttpResponse: Excel文件响应
        """
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # 创建工作簿
        wb = Workbook()
        
        # ========== 概览工作表 ==========
        ws_overview = wb.active
        ws_overview.title = "扫描概览"
        
        # 定义样式
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2d3343", end_color="2d3343", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 获取漏洞数据
        vulnerabilities = Vulnerability.objects.filter(task=scan_task)
        
        # 统计数据
        summary = {
            "total_vulnerabilities": vulnerabilities.count(),
            "critical": vulnerabilities.filter(risk_level="critical").count(),
            "high": vulnerabilities.filter(risk_level="high").count(),
            "medium": vulnerabilities.filter(risk_level="medium").count(),
            "low": vulnerabilities.filter(risk_level="low").count(),
            "info": vulnerabilities.filter(risk_level="info").count(),
        }
        
        # 获取技术栈信息
        technology_stack = {}
        try:
            scan_result = ScanResult.objects.get(task=scan_task)
            technology_stack = scan_result.technology_stack or {}
        except ScanResult.DoesNotExist:
            pass
        
        # 标题
        ws_overview.merge_cells('A1:D1')
        ws_overview['A1'] = "Web应用程序漏洞检测报告"
        ws_overview['A1'].font = Font(bold=True, size=18)
        ws_overview['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        ws_overview['A3'] = "基本信息"
        ws_overview['A3'].font = subheader_font
        ws_overview.merge_cells('A3:D3')
        ws_overview['A3'].fill = subheader_fill
        
        basic_info = [
            ("任务ID", scan_task.task_id),
            ("目标URL", scan_task.target_url),
            ("扫描时间", scan_task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan_task.completed_at else "N/A"),
            ("扫描持续时间", f"{scan_task.duration_seconds()} 秒"),
            ("扫描页面数", scan_task.pages_scanned or 0),
            ("执行模块数", len(scan_task.custom_modules) if scan_task.custom_modules else 5),
        ]
        
        for idx, (label, value) in enumerate(basic_info, start=4):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = str(value)
            ws_overview.merge_cells(f'B{idx}:D{idx}')
        
        # 漏洞统计
        row = len(basic_info) + 5
        ws_overview[f'A{row}'] = "漏洞统计"
        ws_overview[f'A{row}'].font = subheader_font
        ws_overview.merge_cells(f'A{row}:D{row}')
        ws_overview[f'A{row}'].fill = subheader_fill
        
        vuln_stats = [
            ("总漏洞数", summary["total_vulnerabilities"]),
            ("危急", summary["critical"]),
            ("高危", summary["high"]),
            ("中危", summary["medium"]),
            ("低危", summary["low"]),
            ("信息", summary["info"]),
        ]
        
        # 风险等级颜色
        risk_colors = {
            "危急": "dc3545",
            "高危": "fd7e14", 
            "中危": "ffc107",
            "低危": "28a745",
            "信息": "17a2b8",
        }
        
        for idx, (label, value) in enumerate(vuln_stats, start=row+1):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = value
            if label in risk_colors:
                ws_overview[f'A{idx}'].fill = PatternFill(start_color=risk_colors[label], end_color=risk_colors[label], fill_type="solid")
                ws_overview[f'A{idx}'].font = Font(bold=True, color="FFFFFF")
        
        # 技术栈信息
        row = row + len(vuln_stats) + 2
        ws_overview[f'A{row}'] = "技术栈信息"
        ws_overview[f'A{row}'].font = subheader_font
        ws_overview.merge_cells(f'A{row}:D{row}')
        ws_overview[f'A{row}'].fill = subheader_fill
        
        tech_info = [
            ("服务器", technology_stack.get('server', '未知')),
            ("编程语言", technology_stack.get('language', '未知')),
            ("框架", technology_stack.get('framework', '未知')),
            ("数据库", technology_stack.get('database', '未知')),
        ]
        
        for idx, (label, value) in enumerate(tech_info, start=row+1):
            ws_overview[f'A{idx}'] = label
            ws_overview[f'A{idx}'].font = Font(bold=True)
            ws_overview[f'B{idx}'] = str(value)
            ws_overview.merge_cells(f'B{idx}:D{idx}')
        
        # 设置列宽
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 30
        ws_overview.column_dimensions['C'].width = 20
        ws_overview.column_dimensions['D'].width = 20
        
        # ========== 漏洞详情工作表 ==========
        ws_vulns = wb.create_sheet("漏洞详情")
        
        # 表头
        headers = ["序号", "漏洞名称", "风险等级", "CVSS评分", "漏洞类型", "URL", "参数", "请求方法"]
        if include_evidence:
            headers.extend(["攻击载荷", "证据"])
        headers.extend(["漏洞描述", "修复建议", "发现时间"])
        
        for col, header in enumerate(headers, start=1):
            cell = ws_vulns.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 填充漏洞数据
        for idx, vuln in enumerate(vulnerabilities, start=1):
            row = idx + 1
            col = 1
            
            ws_vulns.cell(row=row, column=col, value=idx).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.name).border = border
            col += 1
            
            # 风险等级单元格
            risk_cell = ws_vulns.cell(row=row, column=col, value=vuln.risk_level.upper())
            risk_cell.border = border
            risk_colors_cell = {
                'critical': "dc3545",
                'high': "fd7e14",
                'medium': "ffc107",
                'low': "28a745",
                'info': "17a2b8",
            }
            if vuln.risk_level.lower() in risk_colors_cell:
                risk_cell.fill = PatternFill(start_color=risk_colors_cell[vuln.risk_level.lower()], 
                                            end_color=risk_colors_cell[vuln.risk_level.lower()], 
                                            fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            col += 1
            
            ws_vulns.cell(row=row, column=col, value=vuln.cvss_score).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.type).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.url).border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.parameter or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.method).border = border
            col += 1
            
            if include_evidence:
                ws_vulns.cell(row=row, column=col, value=vuln.payload or "N/A").border = border
                col += 1
                ws_vulns.cell(row=row, column=col, value=vuln.evidence or "N/A").border = border
                col += 1
            
            ws_vulns.cell(row=row, column=col, value=vuln.description or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, value=vuln.remediation or "N/A").border = border
            col += 1
            ws_vulns.cell(row=row, column=col, 
                         value=vuln.detected_at.strftime("%Y-%m-%d %H:%M:%S") if vuln.detected_at else "N/A").border = border
        
        # 设置列宽
        column_widths = [8, 30, 10, 10, 15, 40, 15, 10]
        if include_evidence:
            column_widths.extend([30, 30])
        column_widths.extend([40, 40, 20])
        
        for col, width in enumerate(column_widths, start=1):
            ws_vulns.column_dimensions[get_column_letter(col)].width = width
        
        # ========== 风险统计图表数据工作表 ==========
        ws_stats = wb.create_sheet("统计图表数据")
        
        ws_stats['A1'] = "风险等级"
        ws_stats['B1'] = "数量"
        ws_stats['A1'].font = Font(bold=True)
        ws_stats['B1'].font = Font(bold=True)
        
        stats_data = [
            ("危急", summary["critical"]),
            ("高危", summary["high"]),
            ("中危", summary["medium"]),
            ("低危", summary["low"]),
            ("信息", summary["info"]),
        ]
        
        for idx, (label, value) in enumerate(stats_data, start=2):
            ws_stats[f'A{idx}'] = label
            ws_stats[f'B{idx}'] = value
        
        ws_stats.column_dimensions['A'].width = 15
        ws_stats.column_dimensions['B'].width = 10
        
        # 写入缓冲区
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # 创建HTTP响应
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="scan_report_{scan_task.task_id}.xlsx"'
        
        return response

    def _export_json_report(self, scan_task, include_evidence, include_screenshots):
        """
        导出JSON报告
        
        Args:
            scan_task: 扫描任务对象
            include_evidence: 是否包含攻击证据
            include_screenshots: 是否包含截图信息
            
        Returns:
            HttpResponse: JSON文件响应
        """
        from django.http import HttpResponse
        import json
        
        # 获取漏洞数据
        vulnerabilities = Vulnerability.objects.filter(task=scan_task)
        
        # 统计数据
        summary = {
            "total_vulnerabilities": vulnerabilities.count(),
            "critical": vulnerabilities.filter(risk_level="critical").count(),
            "high": vulnerabilities.filter(risk_level="high").count(),
            "medium": vulnerabilities.filter(risk_level="medium").count(),
            "low": vulnerabilities.filter(risk_level="low").count(),
            "info": vulnerabilities.filter(risk_level="info").count(),
            "pages_scanned": scan_task.pages_scanned,
            "modules_executed": len(scan_task.custom_modules) if scan_task.custom_modules else 5,
        }
        
        # 获取技术栈信息
        technology_stack = {}
        try:
            scan_result = ScanResult.objects.get(task=scan_task)
            technology_stack = scan_result.technology_stack or {}
        except ScanResult.DoesNotExist:
            pass
        
        # 构建漏洞列表
        vuln_list = []
        for vuln in vulnerabilities:
            vuln_data = {
                "id": vuln.id,
                "name": vuln.name,
                "type": vuln.type,
                "risk_level": vuln.risk_level,
                "cvss_score": vuln.cvss_score,
                "cvss_vector": vuln.cvss_vector,
                "url": vuln.url,
                "method": vuln.method,
                "parameter": vuln.parameter,
                "description": vuln.description,
                "remediation": vuln.remediation,
                "detected_at": vuln.detected_at.isoformat() if vuln.detected_at else None,
            }
            
            if include_evidence:
                vuln_data["payload"] = vuln.payload
                vuln_data["evidence"] = vuln.evidence
            
            if include_screenshots and hasattr(vuln, 'screenshots'):
                vuln_data["screenshots"] = vuln.screenshots
            
            if hasattr(vuln, 'references') and vuln.references:
                vuln_data["references"] = vuln.references
            
            vuln_list.append(vuln_data)
        
        # 构建完整报告
        report = {
            "report_info": {
                "task_id": scan_task.task_id,
                "target_url": scan_task.target_url,
                "scan_time": scan_task.completed_at.isoformat() if scan_task.completed_at else None,
                "scan_duration_seconds": scan_task.duration_seconds(),
                "generated_at": timezone.now().isoformat(),
            },
            "summary": summary,
            "technology_stack": technology_stack,
            "vulnerabilities": vuln_list,
        }
        
        # 创建HTTP响应
        response = HttpResponse(
            json.dumps(report, ensure_ascii=False, indent=2),
            content_type="application/json; charset=utf-8"
        )
        response["Content-Disposition"] = f'attachment; filename="scan_report_{scan_task.task_id}.json"'
        
        return response
